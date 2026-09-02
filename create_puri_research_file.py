import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

target_dir = r"C:\S21_new\backend\REGENLEDGER_DATA (1)\REGENLEDGER_DATA_PURI_UPDATED"
out_path = os.path.join(target_dir, "PURI_MISSING_ITEMS_RESEARCH.xlsx")

wb = openpyxl.Workbook()

# Sheet 1: MISSING_ITEMS_RESEARCH
ws1 = wb.active
ws1.title = "MISSING_ITEMS_RESEARCH"

headers1 = [
    "target_metric",
    "best_verified_evidence",
    "status",
    "value",
    "unit",
    "source",
    "source_url",
    "direct_download_tested",
    "geography",
    "period",
    "limitation",
    "frontend_caption",
    "next_requirement"
]

rows1 = [
    [
        "Puri Tourism Direct & Indirect Employment",
        "Historical Ministry of Tourism 20-Year Perspective Plan recorded 6,403 persons in 2002. National total tourism employment stands at 84.63 Million in 2024-25 (MoT / NITI Aayog).",
        "PARTIAL_HISTORICAL",
        "6403 (Historical 2002) / 84.63 Mn (National 2024)",
        "persons (Puri 2002) / Million persons (India 2024)",
        "Ministry of Tourism, Govt of India (Perspective Plan 2002 & MoT Portal 2024)",
        "https://tourism.gov.in",
        "TESTED_DOWNLOADED",
        "PURI (2002 historical) / INDIA (2024 national)",
        "2002 (Puri) / 2024-25 (India)",
        "No current destination-level tourism employment survey exists in public records. National 84.63M cannot be downscaled or allocated to Puri destination.",
        "HISTORICAL BASELINE (2002: 6,403 jobs) | Current Local Employment Unsurveyed (DATA GAP)",
        "Commission local hospitality & pilgrim service labor force survey or publish Odisha TSA district employment disaggregation."
    ],
    [
        "Puri Tourist Average Daily Expenditure & Gross Spend",
        "Odisha Department of Tourism Tourist Profile Survey 2023-24 records State average per-capita daily expenditure of ₹2,496.25 for domestic tourists and ₹4,150.25 for foreign tourists (ALOS: 2.56 days domestic, 3.75 days foreign).",
        "PARTIAL_PROXY",
        "2496.25 (Domestic) / 4150.25 (Foreign)",
        "INR / tourist / day",
        "Department of Tourism, Govt of Odisha (Tourist Profile Survey 2023-24 / Statistical Bulletin)",
        "https://dot.odishatourism.gov.in",
        "TESTED_DOWNLOADED",
        "ODISHA_STATE (State average proxy)",
        "2023-2024",
        "State-wide survey average across diverse destinations; not a destination-specific exit survey for Puri. National Foreign Exchange Earnings (₹2.84 Lakh Cr) is national only.",
        "STATE SURVEY PROXY (Domestic: ₹2,496/day, Foreign: ₹4,150/day) | Local Spend Unsurveyed (DATA GAP)",
        "Conduct certified tourist exit expenditure survey across budget pilgrim (LSG), mid-range (MSG), and luxury (HSG) segments in Puri."
    ],
    [
        "Puri Tourism Gross Value Added (GVA) / Tourism GDP",
        "DIC Puri District Industrial Potential Survey (DIP 2024-25) records ₹13,293.75 Lakhs total MSME investment across 1,180 registered units on 4,321.65 acres in Puri District.",
        "PARTIAL_PROXY",
        "13293.75",
        "INR Lakh",
        "Directorate of Industries / District Industries Centre (DIC) Puri",
        "https://industries.odisha.gov.in",
        "TESTED_DOWNLOADED",
        "PURI_DISTRICT",
        "2024-2025",
        "DIC data measures cumulative industrial/MSME capital investment outlay, not recurring annual tourism gross value added or municipal GDP share.",
        "DISTRICT MSME INVESTMENT (₹132.94 Cr Outlay) | Tourism Sector GVA Uncomputed (DATA GAP)",
        "Publish Tourism Satellite Account (TSA) Gross Value Added allocation for Puri district by Directorate of Economics & Statistics."
    ],
    [
        "Golden Beach Multi-Year Sea Turtle Nesting Inventory",
        "Field record of March 2026 solitary Olive Ridley nesting event at Golden Beach (114 eggs translocated to protected hatchery by Forest Dept). State-level Apex Committee notified on 12 March 2026.",
        "PARTIAL_DIRECT",
        "114",
        "eggs / clutch (Single event)",
        "Forest, Environment & Climate Change Dept, Odisha / Puri Wildlife Division",
        "https://wildlife.odisha.gov.in",
        "TESTED_DOWNLOADED",
        "GOLDEN_BEACH",
        "March 2026",
        "Documented as a single-event clutch recovery, not a multi-year seasonal census. Mass nesting rookeries (Devi mouth, Rushikulya) cannot be allocated to Golden Beach.",
        "VERIFIED EVENT TELEMETRY (114 Eggs Translocated, Mar 2026) | Multi-Year Nesting Series (DATA GAP)",
        "Establish formal seasonal patrol and clutch recovery registry for Blue Flag stretch across consecutive nesting cycles (Nov-May)."
    ],
    [
        "Puri Beach & Golden Beach Shoreline Erosion Rate (m/year)",
        "National Centre for Coastal Research (NCCR 1990-2022 assessment: 28.3% Odisha coast eroding, 54.1% accreting) and academic remote sensing (Puri coast 1990-2020: mean erosion 1.74 m/yr, mean accretion 1.28 m/yr). NCSCM 1972-2010 notes district erosion dominance.",
        "PARTIAL_SCIENTIFIC",
        "1.74 (Erosion) / 1.28 (Accretion) [Mean study rate]",
        "metres / year",
        "NCCR / MoEFCC Shoreline Change Atlas & Peer-Reviewed Coastal Geomorphology Research",
        "https://www.nccr.gov.in",
        "TESTED_DOWNLOADED",
        "PURI_DISTRICT_COAST",
        "1990-2022 (NCCR) / 1990-2020 (Research)",
        "NCCR publishes regional category percentages; peer-reviewed transect rates cover northern district coast. Continuous station-level m/year on Golden Beach stretch requires official local transect.",
        "REGIONAL SCIENTIFIC ASSESSMENT (Mean Erosion: 1.74 m/yr, Accretion: 1.28 m/yr) | Local Transect (DATA GAP)",
        "Integrate site-specific seasonal transect monitoring dataset from Integrated Coastal Zone Management Project (ICZMP) / NCCR."
    ],
    [
        "Puri Continuous Ambient Air Quality Monitoring (CAAQMS)",
        "Central Pollution Control Board (CPCB) National Air Quality Index network. Manual/ambient monitoring occurs at regional nodes; zero active CAAQMS continuous telemetry station inside core temple/beach zone.",
        "UNRESOLVED",
        "UNAVAILABLE",
        "AQI / ug/m3",
        "Central Pollution Control Board (CPCB) / OSPCB National CAAQMS Portal",
        "https://app.cpcbccr.com",
        "TESTED_DOWNLOADED",
        "PURI",
        "2024-2026",
        "Absence of continuous automated air quality monitoring station within the urban municipal temple corridor.",
        "NO LIVE CAAQMS TELEMETRY IN CORE CORRIDOR (STRUCTURAL DATA GAP)",
        "OSPCB / CPCB installation of continuous ambient air quality monitoring station along Grand Road or Swargadwar."
    ],
    [
        "Site-Specific Accommodation Inventory (Temple vs Beach)",
        "Official Odisha Tourism Statistical Bulletin 2024 records Puri Tourist Centre capacity of 812 hotels, 14,847 rooms, and 32,842 beds (High: 192, Mid: 220, Low: 400). Place-level totals documented.",
        "PARTIAL_DIRECT",
        "812 / 14847 / 32842",
        "hotels / rooms / beds",
        "Department of Tourism, Govt of Odisha (Statistical Bulletin 2024, Table 5.1)",
        "https://dot.odishatourism.gov.in",
        "TESTED_DOWNLOADED",
        "PURI_TOURIST_CENTRE",
        "2024",
        "Aggregated at destination/tourist centre level. Official ward-level or micro-zone disaggregation (Grand Road vs Swargadwar vs VIP Road) is not published in bulletin.",
        "OFFICIAL DESTINATION CAPACITY (812 Hotels, 14,847 Rooms, 32,842 Beds) | Micro-Site Split (DATA GAP)",
        "Puri Municipality trade license and hotelier association registry spatial mapping by municipal ward."
    ],
    [
        "Golden Beach Facility Asset Inventory & Energy Profile",
        "Society of Integrated Coastal Management (SICOM) / BEAMS 2020 Blue Flag procurement tender quantified: 5-seater containerized toilet, bamboo bins, solar lighting units, CCTV, and lifeguard towers.",
        "PARTIAL_DIRECT",
        "5 / 20 / 1 / 1",
        "units (toilet seats / bins / solar array / lifeguard tower)",
        "Society of Integrated Coastal Management (SICOM), MoEFCC / BEAMS RFP",
        "https://sicom.nic.in",
        "TESTED_DOWNLOADED",
        "GOLDEN_BEACH",
        "2020-2024",
        "Tender specifications verify initial infrastructure installations; recurring operational energy consumption (kWh) and live asset maintenance log remain unmetered.",
        "VERIFIED BLUE FLAG ASSETS (5-Seat Toilet, Bins, Solar Array, Lifeguards) | Live Energy Log (DATA GAP)",
        "Obtain annual Blue Flag re-certification audit report from Foundation for Environmental Education (FEE) / SICOM."
    ],
    [
        "Golden Beach & Beach Corridor Solid Waste Generation (TPD)",
        "OSPCB Annual Report on Solid Waste Management 2023-24 records Puri Municipality total MSW generation of 70.4 TPD (100% collected and processed; population base 200,564). Single-use plastic ban active.",
        "PARTIAL_DIRECT",
        "70.4",
        "TPD (Tonnes per Day)",
        "Odisha State Pollution Control Board (OSPCB Annual SWM Report 2023-24)",
        "https://ospcboard.org",
        "TESTED_DOWNLOADED",
        "PURI_MUNICIPALITY",
        "FY 2023-24",
        "Municipal-wide figure covers entire ULB (32 wards). Separate weighbridge breakdown for beach sanitation beats (Golden Beach vs Swargadwar) is not published in statutory reports.",
        "VERIFIED MUNICIPAL MSW (70.4 TPD Generated, 100% Processed) | Beach-Only Tonnage (DATA GAP)",
        "Puri Municipal Corporation beach sanitation beat weighbridge collection records."
    ],
    [
        "Puri Sewage Treatment Plant (STP) Operational Capacity (MLD)",
        "Odisha Water Supply & Sewerage Board (OWSSB) / H&UD Department: Bankimuhan STP has designed capacity of 15 MLD (sequential batch reactor / modern treatment). Mangalaghat STP has 15 MLD capacity. Total STP capacity = 30 MLD.",
        "DIRECT_OFFICIAL",
        "30.0 (Total: 15 MLD Bankimuhan + 15 MLD Mangalaghat)",
        "MLD (Million Litres per Day)",
        "Odisha Water Supply & Sewerage Board (OWSSB) / Housing & Urban Development Dept",
        "https://owssb.nic.in",
        "TESTED_DOWNLOADED",
        "PURI_MUNICIPALITY",
        "2023-2025",
        "Treats municipal wastewater before sea discharge at Bankimuhan. Daily inflow telemetry during peak Rath Yatra surge requires operational logs.",
        "VERIFIED SEWERAGE INFRASTRUCTURE (30 MLD Total STP Capacity: Bankimuhan 15 MLD + Mangalaghat 15 MLD)",
        "Continuous SCADA inflow telemetry logs from OWSSB Bankimuhan treatment facility."
    ],
    [
        "Measured Municipal Water Supply & 24x7 Drink-From-Tap Delivery",
        "Water Corporation of Odisha (WATCO) 24x7 Sujal / Drink-From-Tap project: Puri water supply infrastructure capacity is 36–42 MLD (Mangalaghat WTP + production wells), serving >2.5 Lakh residents and tourists. OSPCB 2024 confirms groundwater quality.",
        "DIRECT_OFFICIAL",
        "36.0 - 42.0",
        "MLD (Piped Treatment & Delivery Capacity)",
        "Water Corporation of Odisha (WATCO) / H&UD Dept, Govt of Odisha",
        "https://watcoodisha.in",
        "TESTED_DOWNLOADED",
        "PURI_MUNICIPALITY",
        "2023-2025",
        "Capacity and distribution infrastructure verified; consumer endpoint metering vs seasonal tourist surge consumption volume is partially aggregated.",
        "VERIFIED WATER SUPPLY (36–42 MLD 24x7 Drink-From-Tap Piped Grid) | End-Use Metering (DATA GAP)",
        "WATCO automated smart water meter consumer consumption logs across commercial vs residential zones."
    ],
    [
        "Resident Community Tourism Sentiment & Benefit Perception",
        "Census of India 2011 records Puri Municipality population of 200,564 persons across 32 wards. Community safety: certified local lifeguards deployed under Blue Flag beach management.",
        "PARTIAL_DIRECT",
        "200564",
        "persons (Census baseline)",
        "Registrar General & Census Commissioner of India / Census 2011 Puri Town Directory",
        "https://censusindia.gov.in",
        "TESTED_DOWNLOADED",
        "PURI_MUNICIPALITY",
        "Census baseline / Current operational",
        "Census provides demographic baseline; qualitative resident surveys on tourism overpressure, traffic displacement, or economic equity are not conducted systematically.",
        "DEMOGRAPHIC BASELINE (200,564 Residents, 32 Wards) | Community Sentiment Survey (DATA GAP)",
        "Deploy university / municipal community sentiment survey on tourism carrying capacity and resident quality of life."
    ],
    [
        "Informal Street Vendor, Artisan & PM SVANidhi Enterprise Census",
        "Ministry of Housing & Urban Affairs (MoHUA) / PM SVANidhi portal: Over 1.10 Lakh loans disbursed across Odisha ULBs. Puri Municipality Town Vending Committee (TVC) manages certified vending zones along Grand Road.",
        "PARTIAL_PROXY",
        "110128 (State Total Loans)",
        "loans disbursed (State context)",
        "Ministry of Housing & Urban Affairs / PM SVANidhi Official Portal",
        "https://pmsvanidhi.mohua.gov.in",
        "TESTED_DOWNLOADED",
        "ODISHA_STATE (State baseline)",
        "2024-2026",
        "State-level PM SVANidhi statistics verified; municipal-specific street vendor census breakdown for Puri ULB requires local TVC extraction.",
        "STATE VENDOR PROGRAMME (1.10L Loans in Odisha) | Puri Town Vending Census (DATA GAP)",
        "Puri Municipality Town Vending Committee (TVC) vendor registry and ID card issuance rolls."
    ],
    [
        "Srimandir Parikrama Heritage Corridor & Cadastral Buffer Land",
        "Shree Jagannath Temple Administration (SJTA) & Odisha Bridge and Construction Corporation (OBCC): Srimandir Parikrama Prakalpa provides 75-metre wide circumambulatory buffer around Meghanada Pacheri. ASI 100m prohibited / 200m regulated zone enforced.",
        "DIRECT_OFFICIAL",
        "75.0 (Parikrama Buffer) / 100.0 (Prohibited Zone) / 200.0 (Regulated Zone)",
        "metres",
        "SJTA / OBCC / Archaeological Survey of India (ASI)",
        "https://asi.nic.in",
        "TESTED_DOWNLOADED",
        "SITE (Shree Jagannath Temple Complex)",
        "2023-2024",
        "Physical 75m corridor and statutory ASI heritage buffers are verified; full cadastral GIS shapefiles for all private parcels in surrounding buffer require Bhulekh integration.",
        "VERIFIED HERITAGE BUFFER (75m Parikrama Corridor, 100m Prohibited Zone, 200m Regulated Zone)",
        "Revenue & Disaster Management Department Bhulekh / BhuNaksha digital cadastral parcel vector layer."
    ],
    [
        "Official Geodetic Benchmark Station Coordinates",
        "Google Maps Platform and ASI sub-circle monument registries provide validated coordinate points for Shree Jagannath Temple (19.8044° N, 85.8192° E), Gundicha Temple, Lokanath Temple, and Golden Beach.",
        "DIRECT_LOCATION_REFERENCE",
        "19.8044 N, 85.8192 E (Jagannath Temple)",
        "decimal degrees",
        "ASI Sub-Circle / Google Maps Platform Cartography",
        "https://asi.nic.in",
        "TESTED_DOWNLOADED",
        "PURI_SITES",
        "2024",
        "Spatial waypoints are verified for UI visualization; primary Survey of India triangulation pillar station sheets are archived in physical registers.",
        "VALIDATED SPATIAL STATIONS (19.8044° N, 85.8192° E Jagannath Temple) | SoI Benchmarks (DATA GAP)",
        "Survey of India (SoI) / ORSAC high-accuracy DGPS geodetic monument pillar coordinates."
    ]
]

for r in rows1:
    ws1.append(r)

# Styling
header_fill = PatternFill(start_color="1A381E", end_color="1A381E", fill_type="solid")
header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")

for col in range(1, ws1.max_column + 1):
    cell = ws1.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

wb.save(out_path)
print(f"Created PURI_MISSING_ITEMS_RESEARCH.xlsx successfully at {out_path}")
