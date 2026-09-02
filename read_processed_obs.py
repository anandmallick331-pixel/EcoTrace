import openpyxl
import os
import json

base_dir = r"C:\S21_new\backend\REGENLEDGER_DATA (1)\REGENLEDGER_DATA_PURI_UPDATED"
processed_dir = os.path.join(base_dir, "PROCESSED_DATA")

all_records = []

for root, dirs, files in os.walk(processed_dir):
    for f in sorted(files):
        if f.endswith(".xlsx"):
            domain = f.replace("PURI_", "").replace("_PROCESSED.xlsx", "")
            wb = openpyxl.load_workbook(os.path.join(root, f), data_only=True)
            for sname in wb.sheetnames:
                if not sname.endswith("_METADATA") and sname != "PROCESSING_METADATA":
                    ws = wb[sname]
                    rows = list(ws.iter_rows(values_only=True))
                    if len(rows) > 1:
                        headers = [str(c).strip() if c else "" for c in rows[0]]
                        headers_lower = [h.lower() for h in headers]
                        for r_idx, r in enumerate(rows[1:], start=2):
                            if not any(c is not None and str(c).strip() != "" for c in r):
                                continue
                            row_dict = {headers_lower[i]: r[i] for i in range(min(len(headers_lower), len(r)))}
                            row_dict["_domain"] = domain
                            row_dict["_file"] = f
                            row_dict["_sheet"] = sname
                            row_dict["_orig_row"] = r_idx
                            all_records.append(row_dict)
            wb.close()

print(f"Total raw processed records read: {len(all_records)}")
with open(r"c:\S21_new\puri_raw_obs_inspect.json", "w", encoding="utf-8") as f:
    json.dump(all_records, f, indent=2, default=str)
