import openpyxl
import os

base_dir = r"C:\S21_new\backend\REGENLEDGER_DATA (1)\REGENLEDGER_DATA_PURI_UPDATED"

# 1. Sources in 04_SOURCES.xlsx
wb_s4 = openpyxl.load_workbook(os.path.join(base_dir, "s21_ready", "04_SOURCES.xlsx"), data_only=True)
ws_s4 = wb_s4["SOURCES"]
s4_sources = set([r[0] for r in list(ws_s4.iter_rows(values_only=True))[1:] if r[0]])
wb_s4.close()
print(f"04_SOURCES count: {len(s4_sources)}")

# 2. Sources in METRIC_EVIDENCE_MAP
wb_ev = openpyxl.load_workbook(os.path.join(base_dir, "PROVENANCE_AND_REFERENCES", "PURI_METRIC_EVIDENCE_MAPPING (1).xlsx"), data_only=True)
ws_ev = wb_ev["METRIC_EVIDENCE_MAP"]
ev_sources = set()
for r in list(ws_ev.iter_rows(values_only=True))[1:]:
    s_val = r[15] # primary_source_ids
    if s_val and s_val != "None":
        for s in str(s_val).replace(";", ",").split(","):
            if s.strip():
                ev_sources.add(s.strip())
wb_ev.close()
print(f"METRIC_EVIDENCE_MAP sources count: {len(ev_sources)}")

diff_ev_s4 = ev_sources - s4_sources
print(f"\nSources in Evidence Mapping but not in 04_SOURCES: {len(diff_ev_s4)}")
for s in sorted(diff_ev_s4):
    print("  +", s)

total_union = s4_sources.union(ev_sources)
print(f"\nTotal Unified Verified Sources across entire package: {len(total_union)}")
for s in sorted(total_union):
    print("  ", s)
