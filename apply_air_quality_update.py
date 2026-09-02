import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
import json
import shutil

base_dir = r"C:\S21_new\backend\REGENLEDGER_DATA (1)\REGENLEDGER_DATA_PURI_UPDATED"
framework_dir = os.path.join(base_dir, "framework")
s21_dir = os.path.join(base_dir, "s21_ready")
qa_dir = os.path.join(base_dir, "_QA")

header_fill = PatternFill(start_color="1A381E", end_color="1A381E", fill_type="solid")
header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")

print("========================================================================")
print("UPDATING AIR QUALITY EVIDENCE (OSPCB AR 2019-20) ACROSS PURI REPOSITORY")
print("========================================================================")

# 1. Update 04_SOURCES.xlsx
src_file = os.path.join(s21_dir, "04_SOURCES.xlsx")
wb_s = openpyxl.load_workbook(src_file)
ws_s = wb_s["SOURCES"]

# Check if SRC_OSPCB_AR_2019_20 exists, if not append
existing_srcs = [r[0] for r in ws_s.iter_rows(values_only=True)]
if "SRC_OSPCB_AR_2019_20" not in existing_srcs:
    ws_s.append([
        "SRC_OSPCB_AR_2019_20",
        "Odisha State Pollution Control Board",
        "Annual Report 2019-20",
        "2020",
        "ANNUAL_REPORT",
        "https://ospcboard.org",
        "files_1650444027_2021029350.pdf (pp. 9-10 / iv-v)",
        "PURI",
        "ENVIRONMENT",
        "2026-08-25",
        "Verified official annual report confirming Puri ambient air monitoring network (8 parameters under NAMP/SAMP)."
    ])
    for col in range(1, ws_s.max_column + 1):
        cell = ws_s.cell(row=ws_s.max_row, column=col)
        cell.font = Font(name="Arial", size=10)
    wb_s.save(src_file)
    shutil.copy2(src_file, os.path.join(r"C:\S21_new\backend\s21_ready_puri", "04_SOURCES.xlsx"))
    print("1. 04_SOURCES.xlsx updated with SRC_OSPCB_AR_2019_20.")

# 2. Update PURI_MISSING_ITEMS_RESEARCH.xlsx
mis_file = os.path.join(base_dir, "PURI_MISSING_ITEMS_RESEARCH.xlsx")
wb_mis = openpyxl.load_workbook(mis_file)
ws_mis = wb_mis["MISSING_ITEMS_RESEARCH"]

for row_idx, r in enumerate(ws_mis.iter_rows(values_only=True), start=1):
    if r[0] and "AIR" in str(r[0]).upper():
        ws_mis.cell(row=row_idx, column=2, value="OSPCB Annual Report 2019-20 (files_1650444027_2021029350.pdf, pp. 9-10 / iv-v) confirms periodic ambient air quality monitoring in Puri across 8 parameters under NAMP/SAMP. No live/continuous CAAQMS telemetry station.")
        ws_mis.cell(row=row_idx, column=3, value="PARTIAL")
        ws_mis.cell(row=row_idx, column=4, value="periodic official NAMP/SAMP monitoring in Puri (8 parameters)")
        ws_mis.cell(row=row_idx, column=5, value="DIRECT")
        ws_mis.cell(row=row_idx, column=6, value="Odisha State Pollution Control Board (Annual Report 2019-20)")
        ws_mis.cell(row=row_idx, column=7, value="files_1650444027_2021029350.pdf")
        ws_mis.cell(row=row_idx, column=8, value="LOCALLY_VERIFIED / DOWNLOAD_TESTED")
        ws_mis.cell(row=row_idx, column=9, value="PURI")
        ws_mis.cell(row=row_idx, column=10, value="2019-20")
        ws_mis.cell(row=row_idx, column=11, value="Periodic manual sampling under NAMP/SAMP exists; live automated CAAQMS telemetry is uninstalled.")
        ws_mis.cell(row=row_idx, column=12, value="PARTIAL EVIDENCE — Periodic official air-quality monitoring exists; no verified live/continuous Puri CAAQMS telemetry.")
        ws_mis.cell(row=row_idx, column=13, value="REPORTING_ONLY: Deploy continuous CAAQMS station at Grand Road / Swargadwar for live telemetry.")

wb_mis.save(mis_file)
shutil.copy2(mis_file, os.path.join(r"C:\S21_new\backend", "PURI_MISSING_ITEMS_RESEARCH.xlsx"))
print("2. PURI_MISSING_ITEMS_RESEARCH.xlsx updated with OSPCB AR 2019-20 evidence.")

# 3. Update PURI_METRIC_STATUS_TABLE.xlsx
mst_file = os.path.join(framework_dir, "PURI_METRIC_STATUS_TABLE.xlsx")
wb_mst = openpyxl.load_workbook(mst_file)
ws_mst = wb_mst["METRIC_STATUS_TABLE"]
mst_headers = [str(c).strip() for c in next(ws_mst.iter_rows(values_only=True))]

air_found = False
for row_idx, r in enumerate(ws_mst.iter_rows(values_only=True), start=1):
    if r[0] and ("AIR" in str(r[0]).upper() or "MET_ENV_AIR" in str(r[0]).upper()):
        air_found = True
        ws_mst.cell(row=row_idx, column=5, value="PARTIAL") # target_status
        ws_mst.cell(row=row_idx, column=6, value="periodic official NAMP/SAMP monitoring in Puri") # current_representation
        ws_mst.cell(row=row_idx, column=7, value="DIRECT") # representation_type
        ws_mst.cell(row=row_idx, column=8, value="REPORTING_ONLY") # resolution_gate

if not air_found:
    ws_mst.append([
        "MET_ENV_LIVE_AIR_QUALITY",
        "ENVIRONMENT",
        "Air Quality Monitoring",
        "Live CAAQMS Air Quality Telemetry",
        "PARTIAL",
        "periodic official NAMP/SAMP monitoring in Puri",
        "DIRECT",
        "REPORTING_ONLY",
        "OSPCB Annual Report 2019-20 (pp. 9-10 / iv-v) confirms 8 parameters monitored under NAMP/SAMP in Puri.",
        "SRC_OSPCB_AR_2019_20",
        "EV_PURI_ENV_AIR_001"
    ])
    for col in range(1, ws_mst.max_column + 1):
        cell = ws_mst.cell(row=ws_mst.max_row, column=col)
        cell.font = Font(name="Arial", size=10)

wb_mst.save(mst_file)
shutil.copy2(mst_file, os.path.join(r"C:\S21_new\backend\framework", "PURI_METRIC_STATUS_TABLE.xlsx"))
print("3. PURI_METRIC_STATUS_TABLE.xlsx updated (target_status=PARTIAL, rep=DIRECT, gate=REPORTING_ONLY).")

# 4. Update Gap Framework Files (COMPUTATIONAL_GAP_MATRIX, GAP_MASTER_RECONCILIATION, GAP_RESOLUTION_GATE, PURI_DATA_GAPS)
gap_files = [
    os.path.join(framework_dir, "PURI_COMPUTATIONAL_GAP_MATRIX.xlsx"),
    os.path.join(framework_dir, "PURI_GAP_MASTER_RECONCILIATION.xlsx"),
    os.path.join(framework_dir, "PURI_GAP_RESOLUTION_GATE.xlsx"),
    os.path.join(base_dir, "PURI_DATA_GAPS.xlsx")
]

for gf in gap_files:
    wb_g = openpyxl.load_workbook(gf)
    ws_g = wb_g.active
    # If gap row exists for air, update it; otherwise append/ensure
    for row_idx, r in enumerate(ws_g.iter_rows(values_only=True), start=1):
        if r[0] and ("AIR" in str(r[0]).upper() or "ENV" in str(r[1]).upper() and "AIR" in str(r[3]).upper()):
            # Update columns
            ws_g.cell(row=row_idx, column=5, value="PARTIAL")
            ws_g.cell(row=row_idx, column=6, value="periodic official NAMP/SAMP monitoring in Puri")
            ws_g.cell(row=row_idx, column=7, value="DIRECT")
            if ws_g.max_column >= 17:
                ws_g.cell(row=row_idx, column=17, value="REPORTING_ONLY")
    wb_g.save(gf)
    shutil.copy2(gf, os.path.join(r"C:\S21_new\backend\framework", os.path.basename(gf)))

print("4. Gap framework matrices updated with PARTIAL / DIRECT / REPORTING_ONLY for air quality.")

# 5. Update CRITICAL_GAPS_RESOLUTION.md
crit_file = os.path.join(s21_dir, "CRITICAL_GAPS_RESOLUTION.md")
with open(crit_file, "r", encoding="utf-8") as f:
    crit_text = f.read()

# Replace or add air quality disclosure
air_disclosure = "| **`GAP_ENV_AIR_001`** | Live Air Quality Monitoring | `PARTIAL` | `DIRECT` | **`REPORTING_ONLY`** | PARTIAL EVIDENCE — Periodic official air-quality monitoring exists; no verified live/continuous Puri CAAQMS telemetry. |"
if "GAP_ENV_AIR" not in crit_text:
    crit_text = crit_text.replace(
        "| **`GAP_VIS_001`** | Turnstile Hourly Footfall | `UNRESOLVED` | `CONTEXT_ONLY` | **`REPORTING_ONLY`** | Report annual footfall (8.34M) and Rath Yatra peak as context only. |",
        "| **`GAP_VIS_001`** | Turnstile Hourly Footfall | `UNRESOLVED` | `CONTEXT_ONLY` | **`REPORTING_ONLY`** | Report annual footfall (8.34M) and Rath Yatra peak as context only. |\n" + air_disclosure
    )
    with open(crit_file, "w", encoding="utf-8") as f:
        f.write(crit_text)
    shutil.copy2(crit_file, os.path.join(r"C:\S21_new\backend\s21_ready_puri", "CRITICAL_GAPS_RESOLUTION.md"))

print("5. CRITICAL_GAPS_RESOLUTION.md updated with explicit UI caption.")

# 6. Update 07_DASHBOARD_SUMMARY.xlsx
dash_file = os.path.join(s21_dir, "07_DASHBOARD_SUMMARY.xlsx")
wb_dsh = openpyxl.load_workbook(dash_file)
ws_dsh = wb_dsh["DASHBOARD_SUMMARY"]

existing_cards = [r[0] for r in ws_dsh.iter_rows(values_only=True)]
air_card_row = None
for idx, card in enumerate(existing_cards, start=1):
    if card and "AIR" in str(card).upper():
        air_card_row = idx
        break

air_card_data = [
    "LIVE_AIR_QUALITY_MONITORING",
    "Periodic NAMP/SAMP",
    "8 parameters monitored (OSPCB 2019-20)",
    "live_air_quality_monitoring",
    "DIRECT",
    "HIGH",
    "OSPCB Annual Report 2019-20 confirms periodic monitoring across 8 parameters in Puri. Continuous CAAQMS telemetry station is absent.",
    "PARTIAL EVIDENCE — Periodic official air-quality monitoring exists; no verified live/continuous Puri CAAQMS telemetry."
]

if air_card_row:
    for col_idx, val in enumerate(air_card_data, start=1):
        ws_dsh.cell(row=air_card_row, column=col_idx, value=val)
else:
    ws_dsh.append(air_card_data)

for col in range(1, ws_dsh.max_column + 1):
    cell = ws_dsh.cell(row=ws_dsh.max_row, column=col)
    cell.font = Font(name="Arial", size=10)

wb_dsh.save(dash_file)
shutil.copy2(dash_file, os.path.join(r"C:\S21_new\backend\s21_ready_puri", "07_DASHBOARD_SUMMARY.xlsx"))
print("6. 07_DASHBOARD_SUMMARY.xlsx updated with air quality telemetry card & caption.")

# Re-run QA suite to verify complete integrity
