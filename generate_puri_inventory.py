import os
import openpyxl
import json
from collections import defaultdict, Counter

base_dir = r"C:\S21_new\backend\REGENLEDGER_DATA (1)\REGENLEDGER_DATA_PURI_UPDATED"

all_files_info = []

total_raw_domains = set()
total_observations = 0
obs_by_domain = Counter()
gaps_by_domain = Counter()
prov_counts = Counter()
status_counts = Counter()
confidence_counts = Counter()

unique_sources = set()
unique_records = set()
unique_metrics = set()
unique_datasets = set()
unique_gaps = set()

duplicate_records = defaultdict(list)
duplicate_sources = defaultdict(list)

scaffold_sheets = []
suspicious_files = []

for root, dirs, files in os.walk(base_dir):
    for file in sorted(files):
        rel_path = os.path.relpath(os.path.join(root, file), base_dir)
        full_path = os.path.join(root, file)
        ext = os.path.splitext(file)[1].lower()
        folder = os.path.dirname(rel_path)
        
        file_obj = {
            "rel_path": rel_path,
            "filename": file,
            "ext": ext,
            "folder": folder,
            "sheets": []
        }
        
        if ext in ['.xlsx', '.xls']:
            wb = openpyxl.load_workbook(full_path, data_only=True)
            for sname in wb.sheetnames:
                ws = wb[sname]
                rows = list(ws.iter_rows(values_only=True))
                row_count = len(rows)
                
                if row_count == 0:
                    file_obj["sheets"].append({
                        "sheet_name": sname,
                        "row_count": 0,
                        "col_count": 0,
                        "populated_rows": 0,
                        "headers": [],
                        "is_header_only": False,
                        "is_empty": True,
                        "type": "EMPTY_SHEET"
                    })
                    scaffold_sheets.append((rel_path, sname, "Empty sheet (0 rows)"))
                    continue
                    
                headers = [str(c).strip() if c is not None else "" for c in rows[0]]
                col_count = len(headers)
                data_rows = rows[1:]
                
                populated = 0
                sheet_sources = set()
                sheet_records = set()
                sheet_years = set()
                sheet_units = set()
                sheet_geos = set()
                sheet_provs = set()
                sheet_statuses = set()
                
                headers_lower = [h.lower() for h in headers]
                
                is_processed_domain_sheet = ("PROCESSED_DATA" in folder and not sname.endswith("_METADATA") and sname != "PROCESSING_METADATA")
                
                if is_processed_domain_sheet:
                    domain_name = file.replace("PURI_", "").replace("_PROCESSED.xlsx", "")
                    total_raw_domains.add(domain_name)
                
                for r_idx, r in enumerate(data_rows, start=2):
                    if not any(c is not None and str(c).strip() != "" for c in r):
                        continue
                    populated += 1
                    
                    row_dict = {headers_lower[i]: r[i] for i in range(min(len(headers_lower), len(r)))}
                    
                    # Extract fields
                    rec_id = row_dict.get("record_id") or row_dict.get("obs_id") or row_dict.get("id") or row_dict.get("observation_id") or row_dict.get("master_record_id")
                    if rec_id is not None and str(rec_id).strip() != "":
                        rid = str(rec_id).strip()
                        sheet_records.add(rid)
                        if is_processed_domain_sheet:
                            duplicate_records[rid].append((rel_path, sname, r_idx))
                            unique_records.add(rid)
                            
                    src_id = row_dict.get("source_id") or row_dict.get("source_code") or row_dict.get("source_reference_id") or row_dict.get("source")
                    if src_id is not None and str(src_id).strip() != "":
                        sid = str(src_id).strip()
                        sheet_sources.add(sid)
                        if is_processed_domain_sheet:
                            duplicate_sources[sid].append((rel_path, sname, r_idx))
                            unique_sources.add(sid)
                            
                    metric_id = row_dict.get("metric_code") or row_dict.get("metric_id") or row_dict.get("indicator") or row_dict.get("metric_name")
                    if metric_id is not None and str(metric_id).strip() != "":
                        mid = str(metric_id).strip()
                        if is_processed_domain_sheet:
                            unique_metrics.add(mid)
                            
                    dataset_id = row_dict.get("dataset_id") or row_dict.get("dataset_code") or row_dict.get("dataset")
                    if dataset_id is not None and str(dataset_id).strip() != "":
                        did = str(dataset_id).strip()
                        if is_processed_domain_sheet:
                            unique_datasets.add(did)
                            
                    gap_id = row_dict.get("master_gap_id") or row_dict.get("gap_id")
                    if gap_id is not None and str(gap_id).strip() != "":
                        gid = str(gap_id).strip()
                        if is_processed_domain_sheet:
                            unique_gaps.add(gid)
                            
                    yr = row_dict.get("year") or row_dict.get("period") or row_dict.get("time_period") or row_dict.get("source_period")
                    if yr is not None and str(yr).strip() != "":
                        sheet_years.add(str(yr).strip())
                        
                    unit = row_dict.get("unit") or row_dict.get("raw_unit") or row_dict.get("metric_unit")
                    if unit is not None and str(unit).strip() != "":
                        sheet_units.add(str(unit).strip())
                        
                    geo = row_dict.get("geographic_scope") or row_dict.get("geography") or row_dict.get("location") or row_dict.get("spatial_scope")
                    if geo is not None and str(geo).strip() != "":
                        sheet_geos.add(str(geo).strip())
                        
                    prov = row_dict.get("value_type") or row_dict.get("provenance_type") or row_dict.get("provenance")
                    if prov is not None and str(prov).strip() != "":
                        pval = str(prov).strip().upper()
                        sheet_provs.add(pval)
                        if is_processed_domain_sheet:
                            prov_counts[pval] += 1
                    elif is_processed_domain_sheet:
                        prov_counts["MISSING_PROVENANCE"] += 1
                        
                    st = row_dict.get("verification_status") or row_dict.get("status") or row_dict.get("audit_status")
                    if st is not None and str(st).strip() != "":
                        sval = str(st).strip().upper()
                        sheet_statuses.add(sval)
                        if is_processed_domain_sheet:
                            status_counts[sval] += 1
                            
                    conf = row_dict.get("confidence")
                    if conf is not None and str(conf).strip() != "":
                        cval = str(conf).strip().upper()
                        if is_processed_domain_sheet:
                            confidence_counts[cval] += 1
                            
                    if is_processed_domain_sheet:
                        total_observations += 1
                        obs_by_domain[domain_name] += 1
                        
                        # check if gap
                        avail = str(row_dict.get("data_availability") or "").strip().upper()
                        if avail == "GAP" or "DATA_GAP" in str(prov or "").upper() or row_dict.get("value") == "DATA_GAP" or gap_id is not None:
                            gaps_by_domain[domain_name] += 1
                
                is_scaffold = (populated == 0 and row_count > 0)
                if is_scaffold:
                    scaffold_sheets.append((rel_path, sname, "Header-only (0 data rows)"))
                    
                file_obj["sheets"].append({
                    "sheet_name": sname,
                    "row_count": row_count,
                    "col_count": col_count,
                    "populated_rows": populated,
                    "headers": headers,
                    "is_header_only": is_scaffold,
                    "is_empty": False,
                    "type": "METADATA_SHEET" if "METADATA" in sname else ("OBSERVATION_SHEET" if is_processed_domain_sheet else "REPORT_OR_QA"),
                    "sample_sources": list(sheet_sources)[:5],
                    "sample_records": list(sheet_records)[:5],
                    "record_count": len(sheet_records),
                    "years": sorted(list(sheet_years))[:8],
                    "units": list(sheet_units)[:6],
                    "geographies": list(sheet_geos)[:6],
                    "provenance_types": list(sheet_provs),
                    "statuses": list(sheet_statuses)
                })
            wb.close()
        elif ext == '.md':
            with open(full_path, 'r', encoding='utf-8', errors='ignore') as f:
                lines = f.readlines()
            file_obj["sheets"].append({
                "sheet_name": "MARKDOWN_DOC",
                "row_count": len(lines),
                "col_count": 1,
                "populated_rows": len([l for l in lines if l.strip()]),
                "headers": ["Markdown Text Content"],
                "is_header_only": False,
                "is_empty": False,
                "type": "DOCUMENTATION"
            })
            
        all_files_info.append(file_obj)

# Collect duplicate checks
dups_record_list = {k: v for k, v in duplicate_records.items() if len(v) > 1}
dups_source_list = {k: len(v) for k, v in duplicate_sources.items() if len(v) > 1}

summary_output = {
    "total_files": len(all_files_info),
    "total_raw_domains": len(total_raw_domains),
    "raw_domains_list": sorted(list(total_raw_domains)),
    "total_observations": total_observations,
    "obs_by_domain": dict(obs_by_domain),
    "gaps_by_domain": dict(gaps_by_domain),
    "provenance_distribution": dict(prov_counts),
    "status_distribution": dict(status_counts),
    "confidence_distribution": dict(confidence_counts),
    "unique_records_count": len(unique_records),
    "unique_sources_count": len(unique_sources),
    "unique_metrics_count": len(unique_metrics),
    "unique_datasets_count": len(unique_datasets),
    "unique_gaps_count": len(unique_gaps),
    "duplicate_records_count": len(dups_record_list),
    "duplicate_records": dups_record_list,
    "scaffold_sheets": scaffold_sheets,
    "all_files": all_files_info
}

with open(r"c:\S21_new\puri_full_inventory.json", "w", encoding="utf-8") as f:
    json.dump(summary_output, f, indent=2)

print("Full inventory generated successfully.")
