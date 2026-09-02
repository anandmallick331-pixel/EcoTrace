import openpyxl
import os
import json

base_dir = r"C:\S21_new\backend\REGENLEDGER_DATA (1)\REGENLEDGER_DATA_PURI_UPDATED"

# 1. Inspect Source Register Validation and All Sources in package
src_val_path = os.path.join(base_dir, "_QA", "PURI_SOURCE_REGISTER_VALIDATION.xlsx")
wb_src = openpyxl.load_workbook(src_val_path, data_only=True)
sources_data = []
for sname in wb_src.sheetnames:
    ws = wb_src[sname]
    rows = list(ws.iter_rows(values_only=True))
    sources_data.append({"sheet": sname, "rows": [list(r) for r in rows]})
wb_src.close()

# 2. Inspect Metric Dictionary
dict_path = os.path.join(base_dir, "METHODOLOGY", "REGENLEDGER_METRIC_DICTIONARY (2).xlsx")
wb_dict = openpyxl.load_workbook(dict_path, data_only=True)
metrics_data = []
ws_m = wb_dict['METRIC_DICTIONARY']
m_rows = list(ws_m.iter_rows(values_only=True))
m_headers = m_rows[0]
for r in m_rows[1:]:
    metrics_data.append(dict(zip(m_headers, r)))
wb_dict.close()

# 3. Inspect Provenance and Reference mappings
ev_map_path = os.path.join(base_dir, "PROVENANCE_AND_REFERENCES", "PURI_METRIC_EVIDENCE_MAPPING (1).xlsx")
wb_ev = openpyxl.load_workbook(ev_map_path, data_only=True)
gaps_data = []
ws_g = wb_ev['DATA_GAP_MAPPING']
g_rows = list(ws_g.iter_rows(values_only=True))
g_headers = g_rows[0]
for r in g_rows[1:]:
    gaps_data.append(dict(zip(g_headers, r)))
wb_ev.close()

# 4. Extract all unique sources from PROCESSED_DATA
all_sources_dict = {}
for root, dirs, files in os.walk(os.path.join(base_dir, "PROCESSED_DATA")):
    for f in files:
        if f.endswith(".xlsx"):
            wb = openpyxl.load_workbook(os.path.join(root, f), data_only=True)
            for sname in wb.sheetnames:
                if not sname.endswith("_METADATA") and sname != "PROCESSING_METADATA":
                    ws = wb[sname]
                    rows = list(ws.iter_rows(values_only=True))
                    if len(rows) > 1:
                        headers = [str(c).lower() if c else "" for c in rows[0]]
                        src_idx = -1
                        for idx, h in enumerate(headers):
                            if "source_id" in h or "source_code" in h or h == "source":
                                src_idx = idx
                                break
                        if src_idx != -1:
                            for r in rows[1:]:
                                if src_idx < len(r) and r[src_idx]:
                                    sid = str(r[src_idx]).strip()
                                    if sid not in all_sources_dict:
                                        all_sources_dict[sid] = {
                                            "source_id": sid,
                                            "domains": set(),
                                            "records": [],
                                            "source_types": set(),
                                            "years": set(),
                                            "geos": set()
                                        }
                                    all_sources_dict[sid]["domains"].add(f.replace("PURI_", "").replace("_PROCESSED.xlsx", ""))
                                    # check record_id
                                    rec_idx = 0 # usually first
                                    if r[rec_idx]:
                                        all_sources_dict[sid]["records"].append(str(r[rec_idx]))
                                    # check source_type
                                    st_idx = headers.index("source_type") if "source_type" in headers else -1
                                    if st_idx != -1 and st_idx < len(r) and r[st_idx]:
                                        all_sources_dict[sid]["source_types"].add(str(r[st_idx]))
                                    # year
                                    yr_idx = headers.index("year") if "year" in headers else -1
                                    if yr_idx != -1 and yr_idx < len(r) and r[yr_idx]:
                                        all_sources_dict[sid]["years"].add(str(r[yr_idx]))
                                    # geo
                                    geo_idx = headers.index("geographic_scope") if "geographic_scope" in headers else -1
                                    if geo_idx != -1 and geo_idx < len(r) and r[geo_idx]:
                                        all_sources_dict[sid]["geos"].add(str(r[geo_idx]))
            wb.close()

# Format output
serializable_sources = {k: {
    "source_id": v["source_id"],
    "domains": list(v["domains"]),
    "record_count": len(v["records"]),
    "sample_records": v["records"][:4],
    "source_types": list(v["source_types"]),
    "years": list(v["years"]),
    "geos": list(v["geos"])
} for k, v in all_sources_dict.items()}

with open(r"c:\S21_new\puri_contract_raw.json", "w", encoding="utf-8") as f:
    json.dump({
        "sources_count": len(serializable_sources),
        "sources": serializable_sources,
        "metrics_count": len(metrics_data),
        "metrics": metrics_data,
        "gaps_count": len(gaps_data),
        "gaps": gaps_data
    }, f, indent=2)

print(f"Contract raw extracted: {len(serializable_sources)} sources, {len(metrics_data)} metrics, {len(gaps_data)} gap entries.")
