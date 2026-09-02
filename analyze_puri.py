import os
import openpyxl
import json
from collections import defaultdict, Counter

base_dir = r"C:\S21_new\backend\REGENLEDGER_DATA (1)\REGENLEDGER_DATA_PURI_UPDATED"

detailed_files = []
all_record_ids = []
all_source_ids = []
all_metric_ids = []
all_dataset_ids = []

provenance_counter = Counter()
status_counter = Counter()
domain_obs_counter = Counter()
domain_gap_counter = Counter()

total_observations = 0
total_gaps = 0

scaffold_files = []
suspicious_files = []

for root, dirs, files in os.walk(base_dir):
    for file in sorted(files):
        rel_path = os.path.relpath(os.path.join(root, file), base_dir)
        full_path = os.path.join(root, file)
        ext = os.path.splitext(file)[1].lower()
        folder = os.path.dirname(rel_path)
        
        # Determine domain from filename or folder
        domain = "GENERAL"
        for d in ["BIODIVERSITY", "COMMUNITY", "ECONOMIC", "EMPLOYMENT", "ENVIRONMENT", 
                  "EXPENDITURE", "GIS", "HERITAGE", "LOCAL_BUSINESS", "OWNERSHIP", 
                  "TOURISM", "VISITOR", "WASTE", "WATER"]:
            if d in file.upper():
                domain = d
                break
        if domain == "GENERAL":
            if "FINAL_OUTPUTS" in folder:
                domain = "FINAL_OUTPUTS"
            elif "METHODOLOGY" in folder:
                domain = "METHODOLOGY"
            elif "PROVENANCE" in folder:
                domain = "PROVENANCE"
            elif "README" in folder:
                domain = "DOCUMENTATION"
            elif "RECONCILIATION" in folder:
                domain = "RECONCILIATION"
            elif "QA" in folder:
                domain = "QA"
            elif "SUPPORTING_METHODOLOGY" in folder:
                domain = "SUPPORTING_METHODOLOGY"
                
        file_summary = {
            "rel_path": rel_path,
            "filename": file,
            "ext": ext,
            "folder": folder,
            "domain": domain,
            "sheets": []
        }
        
        if ext in ['.xlsx', '.xls']:
            wb = openpyxl.load_workbook(full_path, data_only=True)
            for sheetname in wb.sheetnames:
                ws = wb[sheetname]
                rows = list(ws.iter_rows(values_only=True))
                row_count = len(rows)
                
                if row_count == 0:
                    file_summary["sheets"].append({
                        "sheetname": sheetname,
                        "row_count": 0,
                        "col_count": 0,
                        "populated_rows": 0,
                        "headers": [],
                        "is_header_only": False,
                        "is_empty": True,
                        "is_scaffold": True,
                        "notes": "Empty sheet"
                    })
                    scaffold_files.append(f"{rel_path} [{sheetname}] (Empty)")
                    continue
                    
                headers = [str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(rows[0])]
                col_count = len(headers)
                
                # Check rows
                data_rows = rows[1:]
                populated_count = 0
                
                sheet_record_ids = []
                sheet_source_ids = []
                sheet_years = set()
                sheet_units = set()
                sheet_geos = set()
                sheet_provs = set()
                sheet_statuses = set()
                sheet_missing_required = 0
                
                header_lower = [h.lower() for h in headers]
                
                # Find indices of key columns
                def find_col_idx(keys):
                    for k in keys:
                        for idx, h in enumerate(header_lower):
                            if k == h or k in h:
                                return idx
                    return -1
                
                rec_idx = find_col_idx(["record_id", "observation_id", "obs_id", "id", "metric_id"])
                src_idx = find_col_idx(["source_id", "source_code", "source_reference_id", "source"])
                metric_idx = find_col_idx(["metric_code", "metric_id", "metric_name", "metric"])
                dataset_idx = find_col_idx(["dataset_id", "dataset_code", "dataset"])
                year_idx = find_col_idx(["year", "period", "period_start", "time_period", "date"])
                unit_idx = find_col_idx(["unit", "raw_unit", "metric_unit", "standard_unit"])
                geo_idx = find_col_idx(["geography", "geographic_scope", "location", "spatial_scope", "location_name"])
                prov_idx = find_col_idx(["value_type", "provenance_type", "provenance", "methodology_type", "data_type"])
                status_idx = find_col_idx(["status", "verification_status", "quality_status", "review_status"])
                val_idx = find_col_idx(["raw_value", "normalized_value", "value", "reported_value"])
                
                is_obs_sheet = "PROCESSED_DATA" in folder or "RECONCILIATION" in folder or "FINAL_OUTPUTS" in folder
                
                for r in data_rows:
                    if not any(c is not None and str(c).strip() != "" for c in r):
                        continue
                    populated_count += 1
                    
                    row_dict = {headers[i]: r[i] for i in range(min(len(headers), len(r)))}
                    
                    # Extract fields
                    if rec_idx != -1 and rec_idx < len(r) and r[rec_idx] is not None:
                        rid = str(r[rec_idx]).strip()
                        if rid:
                            sheet_record_ids.append(rid)
                            all_record_ids.append(rid)
                            
                    if src_idx != -1 and src_idx < len(r) and r[src_idx] is not None:
                        sid = str(r[src_idx]).strip()
                        if sid:
                            sheet_source_ids.append(sid)
                            all_source_ids.append(sid)
                            
                    if metric_idx != -1 and metric_idx < len(r) and r[metric_idx] is not None:
                        mid = str(r[metric_idx]).strip()
                        if mid:
                            all_metric_ids.append(mid)
                            
                    if dataset_idx != -1 and dataset_idx < len(r) and r[dataset_idx] is not None:
                        did = str(r[dataset_idx]).strip()
                        if did:
                            all_dataset_ids.append(did)
                            
                    if year_idx != -1 and year_idx < len(r) and r[year_idx] is not None:
                        y = str(r[year_idx]).strip()
                        if y:
                            sheet_years.add(y)
                            
                    if unit_idx != -1 and unit_idx < len(r) and r[unit_idx] is not None:
                        u = str(r[unit_idx]).strip()
                        if u:
                            sheet_units.add(u)
                            
                    if geo_idx != -1 and geo_idx < len(r) and r[geo_idx] is not None:
                        g = str(r[geo_idx]).strip()
                        if g:
                            sheet_geos.add(g)
                            
                    if prov_idx != -1 and prov_idx < len(r) and r[prov_idx] is not None:
                        p = str(r[prov_idx]).strip().upper()
                        if p:
                            sheet_provs.add(p)
                            if "PROCESSED_DATA" in folder:
                                provenance_counter[p] += 1
                                
                    if status_idx != -1 and status_idx < len(r) and r[status_idx] is not None:
                        st = str(r[status_idx]).strip().upper()
                        if st:
                            sheet_statuses.add(st)
                            if "PROCESSED_DATA" in folder:
                                status_counter[st] += 1
                                
                    # Check for DATA_GAP
                    is_gap = False
                    for c in r:
                        if c is not None and "DATA_GAP" in str(c).upper():
                            is_gap = True
                            break
                    if is_gap:
                        domain_gap_counter[domain] += 1
                        total_gaps += 1
                        
                    if "PROCESSED_DATA" in folder:
                        total_observations += 1
                        domain_obs_counter[domain] += 1
                
                is_scaffold = (populated_count == 0 and row_count > 0)
                if is_scaffold:
                    scaffold_files.append(f"{rel_path} [{sheetname}] (Header-only, 0 rows)")
                    
                file_summary["sheets"].append({
                    "sheetname": sheetname,
                    "row_count": row_count,
                    "col_count": col_count,
                    "populated_rows": populated_count,
                    "headers": headers,
                    "is_header_only": is_scaffold,
                    "is_empty": False,
                    "is_scaffold": is_scaffold,
                    "sample_source_ids": list(set(sheet_source_ids))[:5],
                    "sample_record_ids": sheet_record_ids[:5],
                    "record_count": len(sheet_record_ids),
                    "years": sorted(list(sheet_years))[:8],
                    "units": list(sheet_units)[:6],
                    "geographies": list(sheet_geos)[:6],
                    "provenance_types": list(sheet_provs),
                    "statuses": list(sheet_statuses)
                })
            wb.close()
        elif ext == '.md':
            file_summary["sheets"].append({
                "sheetname": "MARKDOWN_FILE",
                "row_count": file_summary.get("md_lines", 0),
                "col_count": 1,
                "populated_rows": file_summary.get("md_lines", 0),
                "headers": ["Markdown Document"],
                "is_header_only": False,
                "is_empty": False,
                "is_scaffold": False,
                "notes": "Text Markdown documentation"
            })
            
        detailed_files.append(file_summary)

output = {
    "total_files": len(detailed_files),
    "total_observations_processed_data": total_observations,
    "total_gaps_flagged": total_gaps,
    "domain_observations": dict(domain_obs_counter),
    "domain_gaps": dict(domain_gap_counter),
    "provenance_distribution": dict(provenance_counter),
    "status_distribution": dict(status_counter),
    "unique_record_ids_count": len(set(all_record_ids)),
    "total_record_ids_extracted": len(all_record_ids),
    "unique_source_ids_count": len(set(all_source_ids)),
    "unique_metric_ids_count": len(set(all_metric_ids)),
    "unique_dataset_ids_count": len(set(all_dataset_ids)),
    "scaffold_or_empty_sheets": scaffold_files,
    "files": detailed_files
}

with open(r"c:\S21_new\puri_analysis.json", "w", encoding="utf-8") as f:
    json.dump(output, f, indent=2)

print("Analysis completed successfully.")
print(f"Total files: {len(detailed_files)}")
print(f"PROCESSED_DATA total observations: {total_observations}")
print(f"Unique Source IDs: {len(set(all_source_ids))}")
print(f"Unique Record IDs: {len(set(all_record_ids))}")
