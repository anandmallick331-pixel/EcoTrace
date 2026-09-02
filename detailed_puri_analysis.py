import os
import openpyxl
import json
from collections import defaultdict, Counter

base_dir = r"C:\S21_new\backend\REGENLEDGER_DATA (1)\REGENLEDGER_DATA_PURI_UPDATED"

processed_data_files = []
final_outputs_files = []
methodology_files = []
provenance_files = []
readme_files = []
reconciliation_files = []
supporting_methodology_files = []
qa_files = []

# Detailed tracking
raw_domains = set()
domain_breakdown = defaultdict(lambda: {
    "total_rows": 0,
    "direct": 0,
    "derived": 0,
    "estimated": 0,
    "proxy": 0,
    "inferred": 0,
    "norm": 0,
    "data_gap": 0,
    "unknown": 0,
    "verified": 0,
    "raw_status": 0,
    "flagged": 0,
    "records": set(),
    "sources": set(),
    "metrics": set(),
    "years": set(),
    "units": set(),
    "geos": set(),
    "missing_fields_count": 0
})

all_processed_records = []
all_processed_sources = []
all_processed_metrics = []

header_only_sheets = []
suspicious_files_list = []

for root, dirs, files in os.walk(base_dir):
    for file in sorted(files):
        rel_path = os.path.relpath(os.path.join(root, file), base_dir)
        full_path = os.path.join(root, file)
        ext = os.path.splitext(file)[1].lower()
        folder = os.path.dirname(rel_path)
        
        if "PROCESSED_DATA" in folder:
            domain_name = file.replace("PURI_", "").replace("_PROCESSED.xlsx", "")
            raw_domains.add(domain_name)
            
            wb = openpyxl.load_workbook(full_path, data_only=True)
            for sname in wb.sheetnames:
                ws = wb[sname]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) <= 1:
                    header_only_sheets.append((rel_path, sname, len(rows)))
                    continue
                headers = [str(c).strip() if c is not None else "" for c in rows[0]]
                headers_lower = [h.lower() for h in headers]
                
                # Check for standard fields
                for r_idx, r in enumerate(rows[1:], start=2):
                    if not any(c is not None and str(c).strip() != "" for c in r):
                        continue
                    row_map = {headers_lower[i]: r[i] for i in range(min(len(headers_lower), len(r)))}
                    
                    domain_breakdown[domain_name]["total_rows"] += 1
                    
                    # Record ID
                    rec_id = row_map.get("record_id") or row_map.get("obs_id") or row_map.get("id") or row_map.get("observation_id")
                    if rec_id:
                        all_processed_records.append((str(rec_id).strip(), rel_path, sname, r_idx))
                        domain_breakdown[domain_name]["records"].add(str(rec_id).strip())
                    else:
                        domain_breakdown[domain_name]["missing_fields_count"] += 1
                        
                    # Source ID
                    src_id = row_map.get("source_id") or row_map.get("source_code") or row_map.get("source_reference_id") or row_map.get("source")
                    if src_id:
                        all_processed_sources.append((str(src_id).strip(), rel_path, sname, r_idx))
                        domain_breakdown[domain_name]["sources"].add(str(src_id).strip())
                        
                    # Metric Code
                    m_code = row_map.get("metric_code") or row_map.get("metric_id") or row_map.get("metric_name")
                    if m_code:
                        all_processed_metrics.append((str(m_code).strip(), rel_path))
                        domain_breakdown[domain_name]["metrics"].add(str(m_code).strip())
                        
                    # Years
                    yr = row_map.get("year") or row_map.get("period") or row_map.get("time_period") or row_map.get("period_start")
                    if yr:
                        domain_breakdown[domain_name]["years"].add(str(yr).strip())
                        
                    # Units
                    unit = row_map.get("unit") or row_map.get("raw_unit") or row_map.get("metric_unit")
                    if unit:
                        domain_breakdown[domain_name]["units"].add(str(unit).strip())
                        
                    # Geography
                    geo = row_map.get("geography") or row_map.get("geographic_scope") or row_map.get("location") or row_map.get("spatial_scope")
                    if geo:
                        domain_breakdown[domain_name]["geos"].add(str(geo).strip())
                        
                    # Provenance / Value Type
                    prov = str(row_map.get("value_type") or row_map.get("provenance_type") or row_map.get("provenance") or "").strip().upper()
                    
                    # Check if gap in any column
                    is_gap = ("DATA_GAP" in prov) or any(c is not None and "DATA_GAP" in str(c).upper() for c in r)
                    if is_gap or "GAP" in prov:
                        domain_breakdown[domain_name]["data_gap"] += 1
                    elif "DIRECT" in prov:
                        domain_breakdown[domain_name]["direct"] += 1
                    elif "DERIVED" in prov:
                        domain_breakdown[domain_name]["derived"] += 1
                    elif "ESTIMATED" in prov or "ESTIMATE" in prov:
                        domain_breakdown[domain_name]["estimated"] += 1
                    elif "PROXY" in prov:
                        domain_breakdown[domain_name]["proxy"] += 1
                    elif "INFERRED" in prov:
                        domain_breakdown[domain_name]["inferred"] += 1
                    elif "NORM" in prov:
                        domain_breakdown[domain_name]["norm"] += 1
                    else:
                        domain_breakdown[domain_name]["unknown"] += 1
                        
                    # Status
                    st = str(row_map.get("status") or row_map.get("verification_status") or "").strip().upper()
                    if "VERIFIED" in st:
                        domain_breakdown[domain_name]["verified"] += 1
                    elif "RAW" in st:
                        domain_breakdown[domain_name]["raw_status"] += 1
                    elif "FLAGGED" in st:
                        domain_breakdown[domain_name]["flagged"] += 1
            wb.close()

# Duplicates check
rec_counts = Counter([r[0] for r in all_processed_records])
dup_records = {k: v for k, v in rec_counts.items() if v > 1}

src_counts = Counter([s[0] for s in all_processed_sources])
dup_sources = {k: v for k, v in src_counts.items() if v > 1}

with open(r"c:\S21_new\puri_detailed_summary.json", "w", encoding="utf-8") as f:
    json.dump({
        "raw_domains": list(raw_domains),
        "domain_stats": {k: {
            "total_rows": v["total_rows"],
            "direct": v["direct"],
            "derived": v["derived"],
            "estimated": v["estimated"],
            "proxy": v["proxy"],
            "inferred": v["inferred"],
            "norm": v["norm"],
            "data_gap": v["data_gap"],
            "unknown": v["unknown"],
            "verified": v["verified"],
            "raw_status": v["raw_status"],
            "flagged": v["flagged"],
            "unique_records": len(v["records"]),
            "unique_sources": len(v["sources"]),
            "unique_metrics": len(v["metrics"]),
            "years": sorted(list(v["years"])),
            "units": list(v["units"]),
            "geos": list(v["geos"]),
            "missing_fields": v["missing_fields_count"]
        } for k, v in domain_breakdown.items()},
        "dup_records_count": len(dup_records),
        "sample_dup_records": list(dup_records.items())[:10],
        "dup_sources_count": len(dup_sources),
        "header_only_sheets": header_only_sheets
    }, f, indent=2)

print("Detailed breakdown complete.")
