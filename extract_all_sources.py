import openpyxl
import os
import json

base_dir = r"C:\S21_new\backend\REGENLEDGER_DATA (1)\REGENLEDGER_DATA_PURI_UPDATED"

sources = {}

for root, dirs, files in os.walk(os.path.join(base_dir, "PROCESSED_DATA")):
    for f in files:
        if f.endswith(".xlsx"):
            domain = f.replace("PURI_", "").replace("_PROCESSED.xlsx", "")
            wb = openpyxl.load_workbook(os.path.join(root, f), data_only=True)
            for sname in wb.sheetnames:
                if not sname.endswith("_METADATA") and sname != "PROCESSING_METADATA":
                    ws = wb[sname]
                    rows = list(ws.iter_rows(values_only=True))
                    if len(rows) > 1:
                        headers = [str(c).lower() if c else "" for c in rows[0]]
                        
                        src_idx = headers.index("source_id") if "source_id" in headers else -1
                        type_idx = headers.index("source_type") if "source_type" in headers else -1
                        year_idx = headers.index("year") if "year" in headers else -1
                        geo_idx = headers.index("geographic_scope") if "geographic_scope" in headers else -1
                        rec_idx = headers.index("record_id") if "record_id" in headers else 0
                        notes_idx = headers.index("notes") if "notes" in headers else -1
                        
                        for r in rows[1:]:
                            if src_idx != -1 and r[src_idx]:
                                sid = str(r[src_idx]).strip()
                                if sid not in sources:
                                    sources[sid] = {
                                        "source_id": sid,
                                        "source_type": str(r[type_idx]).strip() if type_idx != -1 and r[type_idx] else "OFFICIAL_PUBLICATION",
                                        "domain": domain,
                                        "domains": set(),
                                        "years": set(),
                                        "geographies": set(),
                                        "records": [],
                                        "notes": []
                                    }
                                sources[sid]["domains"].add(domain)
                                if year_idx != -1 and r[year_idx]:
                                    sources[sid]["years"].add(str(r[year_idx]))
                                if geo_idx != -1 and r[geo_idx]:
                                    sources[sid]["geographies"].add(str(r[geo_idx]))
                                if rec_idx < len(r) and r[rec_idx]:
                                    sources[sid]["records"].append(str(r[rec_idx]))
                                if notes_idx != -1 and r[notes_idx]:
                                    sources[sid]["notes"].append(str(r[notes_idx])[:150])
            wb.close()

print(f"Total Unique Sources Extracted from PROCESSED_DATA: {len(sources)}")
for sid, s in sorted(sources.items()):
    print(f"\n{sid}:")
    print(f"  Type: {s['source_type']} | Domains: {list(s['domains'])} | Years: {list(s['years'])}")
    print(f"  Geographies: {list(s['geographies'])}")
    print(f"  Records ({len(s['records'])}): {s['records'][:3]}")
    if s['notes']:
        print(f"  Sample Note: {s['notes'][0]}")

with open(r"c:\S21_new\puri_sources_extracted.json", "w", encoding="utf-8") as f:
    json.dump({k: {
        "source_id": v["source_id"],
        "source_type": v["source_type"],
        "domains": list(v["domains"]),
        "years": list(v["years"]),
        "geographies": list(v["geographies"]),
        "record_count": len(v["records"]),
        "records": v["records"],
        "sample_note": v["notes"][0] if v["notes"] else ""
    } for k, v in sources.items()}, f, indent=2)
