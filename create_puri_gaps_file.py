import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

target_dir = r"C:\S21_new\backend\REGENLEDGER_DATA (1)\REGENLEDGER_DATA_PURI_UPDATED"
out_path = os.path.join(target_dir, "PURI_DATA_GAPS.xlsx")

wb = openpyxl.Workbook()

# Sheet 1: DATA_GAPS_REGISTER
ws1 = wb.active
ws1.title = "DATA_GAPS_REGISTER"

headers1 = [
    "gap_id",
    "target_metric",
    "description",
    "geography",
    "period",
    "priority",
    "target_status",
    "current_representation",
    "representation_type",
    "source_ids",
    "related_record_ids",
    "next_requirement"
]

rows1 = [
    [
        "GAP_EMP_001",
        "Puri Destination Tourism Employment (Direct + Indirect)",
        "Absence of current destination-level tourism employment series from Ministry of Tourism or Odisha DOT. Historical 2002 figure (6,403 persons) is retained as context but cannot be presented as current.",
        "PURI",
        "2023-2025 / Current",
        "HIGH",
        "UNRESOLVED",
        "Historical 2002 baseline (6,403 persons) documented; National tourism employment (84.63 Mn) documented as national context only",
        "CONTEXT_ONLY",
        "SRC_MOT_DATA_PORTAL, SRC_MOT_20YR_ORISSA_2002",
        "EMP_PURI_HIST_6403_2002, EMP_PURI_HIST_2002, EMP_GAP_CURRENT_PURI, EMP_NATIONAL_CONTEXT",
        "Conduct local employer/hospitality establishment labor census or obtain published Odisha Tourism Satellite Account (TSA) district breakdown."
    ],
    [
        "GAP_EXP_001",
        "Puri Tourist Average Daily Expenditure & Total Spend",
        "No official destination-level tourist expenditure, per-day spend, or shopping spend survey published for Puri. National Foreign Exchange Earnings (FEE ₹2.84 Lakh Cr) cannot be allocated.",
        "PURI",
        "2023-2025 / Current",
        "HIGH",
        "UNRESOLVED",
        "National FEE total (₹2.84 Lakh Cr) documented as national context only; destination spend remains explicit gap",
        "CONTEXT_ONLY",
        "SRC_MOT_DATA_PORTAL",
        "EXP_GAP_PURI_TOURIST_SPEND, EXP_NATIONAL_FEE_CONTEXT",
        "Commission certified tourist exit survey across domestic and international pilgrim/leisure segments in Puri."
    ],
    [
        "GAP_ECO_001",
        "Puri Tourism Gross Value Added (GVA) / Tourism GDP Share",
        "No official tourism-specific GVA or municipal GDP calculation published for Puri destination. DIC district investment pipeline (₹13,293.75 Lakh) reflects investment, not annual tourism GVA.",
        "PURI / PURI_DISTRICT",
        "2023-2025 / Current",
        "MEDIUM",
        "PARTIAL",
        "Puri District MSME total investment and enterprise count documented as district economic baseline",
        "PROXY",
        "SRC_DIC_PURI_DIP_2025",
        "ECO_MSME_INV_2024_25, ECO_MSME_UNITS_2024_25, ECO_MSME_LAND_2024_25, ECO_GAP_TOURISM_GVA",
        "Formal TSA GVA allocation by Directorate of Economics & Statistics, Odisha."
    ],
    [
        "GAP_BIO_001",
        "Golden Beach Multi-Year Sea Turtle Nesting Inventory",
        "Only one-off March 2026 nesting event (114 eggs translocated) documented for Golden Beach. Multi-year seasonal nesting counts are recorded at Devi mouth and Rushikulya rookeries, not Golden Beach.",
        "GOLDEN_BEACH",
        "Multi-year seasonal",
        "MEDIUM",
        "PARTIAL",
        "March 2026 single-clutch nesting event (114 eggs) and state-level protection notification documented",
        "DIRECT",
        "SRC_MEDIA_GB_NESTING_2026, SRC_FECC_TURTLE_COORD_2026, SRC_OMFRA_FISHBAN_2025",
        "BIO_GB_NESTING_2026_03, BIO_PURI_CAMPS_2026, BIO_GAP_GB_MULTI_YEAR",
        "Establish formal Wildlife Division Golden Beach patrol log across consecutive nesting seasons (Nov-May)."
    ],
    [
        "GAP_ENV_001",
        "Puri Beach & Golden Beach Numerical Shoreline Erosion Rate (m/year)",
        "NCSCM 1972-2010 district-level assessment identifies erosion vulnerability across Puri coast, but site-specific numerical erosion rates (metres/year) for Blue Flag stretch are unavailable.",
        "GOLDEN_BEACH / PURI_BEACH",
        "Current / Recent",
        "MEDIUM",
        "PARTIAL",
        "Long-term district shoreline condition (erosion-dominated) and Chilika mouth distance documented",
        "CONTEXT_ONLY",
        "SRC_NCSCM_ODISHA_SHORELINE_2011",
        "ENV_SHORELINE_PURI_DIST_1972_2010, ENV_SHORELINE_CHILIKA_MOUTH, ENV_GAP_SHORELINE_RATE",
        "Obtain high-resolution multi-temporal satellite shoreline transect analysis from NCCR / NCSCM."
    ],
    [
        "GAP_ENV_002",
        "Puri Destination Continuous Ambient Air Quality Monitoring (CAAQMS)",
        "No CPCB or OSPCB continuous ambient air quality monitoring station (CAAQMS) is explicitly mapped to the core temple or beach corridor.",
        "PURI",
        "2024-2025 / Current",
        "LOW",
        "UNRESOLVED",
        "No air quality observation included; gap tracked at domain level",
        "STRUCTURAL_NOT_APPLICABLE",
        "SRC_CPCB_AQI_PORTAL",
        "ENV_GAP_AIR_QUALITY",
        "OSPCB/CPCB installation of CAAQMS station in Puri Grand Road / Swargadwar corridor."
    ],
    [
        "GAP_TOUR_001",
        "Site-Specific Accommodation Inventory (Temple Vicinity vs Beach vs Swargadwar)",
        "Official Odisha Tourism Statistical Bulletins provide destination-wide Puri Tourist Centre capacity (812 hotels, 14,847 rooms, 32,842 beds), but micro-spatial inventory by site/ward is unavailable.",
        "SELECTED_SITES / CORRIDORS",
        "2023-2024",
        "MEDIUM",
        "PARTIAL",
        "Destination-wide category breakdown (HSG 192, MSG 220, LSG 400) and place-level totals documented",
        "DIRECT",
        "SRC_DOT_STAT_BULLETIN_2023, SRC_DOT_STAT_BULLETIN_2024",
        "TOUR_001, TOUR_002, TOUR_003, TOUR_004, TOUR_005, TOUR_GAP_SITE_HOTEL_SPLIT",
        "Municipal trade license / hoteliers association spatial registry by municipal ward."
    ],
    [
        "GAP_TOUR_002",
        "Comprehensive Golden Beach Facility Asset Inventory",
        "2020 Blue Flag RFP tender lists initial amenity specifications (toilets, bamboo huts, bins, solar), but complete post-operational asset count and energy rating remain partial.",
        "GOLDEN_BEACH",
        "Operational Current",
        "MEDIUM",
        "PARTIAL",
        "RFP tender items (5-seater toilet, bamboo bins, solar panels, CCTV, lifeguards) documented",
        "DIRECT",
        "SRC_BEAMS_BLUE_FLAG_RFP_2020",
        "TOUR_GB_AMENITY_TOILETS, TOUR_GB_AMENITY_BINS, TOUR_GB_SOLAR, TOUR_GAP_GB_FULL_INV",
        "Obtain annual Blue Flag compliance and infrastructure audit report from SICOM / MoEFCC."
    ],
    [
        "GAP_WASTE_001",
        "Golden Beach & Swargadwar Beach-Specific Solid Waste Generation (TPD)",
        "OSPCB SWM Annual Reports publish municipal ULB total MSW generation (70.4 TPD in 2023-24, 62 TPD in 2022-23). Beach-specific waste tonnages are not disaggregated.",
        "GOLDEN_BEACH / SWARGADWAR",
        "2023-2024",
        "MEDIUM",
        "PARTIAL",
        "Municipal-wide MSW generation (70.4 TPD), 100% collection, and plastic ban compliance documented",
        "DIRECT",
        "SRC_OSPCB_SWM_2023_24, SRC_OSPCB_PWM_2023_24",
        "WASTE_MSW_GEN_2023_24, WASTE_MSW_COLL_2023_24, WASTE_GAP_BEACH_TPD",
        "Puri Municipality waste weighbridge logs disaggregated by beach sanitation beats."
    ],
    [
        "GAP_WASTE_002",
        "Puri Sewage Treatment Plant (STP) Operational Capacity and Inflow Volume",
        "No official published STP capacity (MLD), treated effluent volume, or sewage network coverage percentage verified in available OSPCB SWM reports.",
        "PURI_MUNICIPALITY",
        "2023-2025 / Current",
        "MEDIUM",
        "UNRESOLVED",
        "No STP numerical capacity row included; tracked as explicit domain infrastructure gap",
        "STRUCTURAL_NOT_APPLICABLE",
        "SRC_WATCO_PURI, SRC_OSPCB_SWM_2023_24",
        "WASTE_GAP_STP_CAPACITY",
        "WATCO / OWSSB published STP operational audit data for Banki Muhana / Puri STPs."
    ],
    [
        "GAP_WAT_001",
        "Puri Measured Municipal & Tourism Water Supply / Consumption (MLD)",
        "OSPCB coastal and groundwater reports provide comprehensive water quality testing, but daily water supply delivery and seasonal tourist consumption volumes are not measured directly.",
        "SITE / PURI_TOURIST_CENTRE",
        "2023-2025 / Current",
        "MEDIUM",
        "PARTIAL",
        "Derived water demand calculation (10.65 MLD) documented based on standard CPHEEO per-capita norms",
        "DERIVED",
        "SRC_OSPCB_GW_2024, SRC_OSPCB_COASTAL_2023",
        "WAT-038, WAT-039, WAT_GAP_MEASURED_CONSUMPTION",
        "WATCO continuous telemetry water meter inflow logs for Puri urban water grid."
    ],
    [
        "GAP_COMM_001",
        "Puri Resident Community Tourism Sentiment & Benefit Perception Survey",
        "Census 2011 provides municipal population (200,564), but empirical community surveys on tourism overpressure, displacement, or cultural sentiment are unpublished.",
        "PURI_MUNICIPALITY",
        "Current",
        "MEDIUM",
        "PARTIAL",
        "Census baseline population and Blue Flag community safety lifeguards documented",
        "DIRECT",
        "SRC_CENSUS_2011_PURI",
        "COMM_POP_2011_CENSUS, COMM_GAP_RESIDENT_SURVEY",
        "Deploy standardized regenerative tourism community benefit & social carrying capacity survey."
    ],
    [
        "GAP_LB_001",
        "Puri Informal Vendor, Handicraft & Temple Artisan Business Registry",
        "Puri has substantial informal street food, souvenir, and handicraft activity around Grand Road, but no formal municipal vendor census is published.",
        "PURI_MUNICIPALITY",
        "Current",
        "MEDIUM",
        "UNRESOLVED",
        "Explicit placeholder record documenting absence of certified street vendor registry",
        "CONTEXT_ONLY",
        "SRC_PURI_MUNICIPALITY",
        "LB_GAP_VENDOR_REGISTRY",
        "Puri Municipality / PM SVANidhi town vending committee survey dataset."
    ],
    [
        "GAP_OWN_001",
        "Cadastral Parcel GIS Mapping of Private vs Municipal vs Endowment Lands",
        "Temple administration statutory ownership (Shree Jagannath Temple Act 1955) is legally verified, but GIS cadastral vector layers for all buffer lands remain unmapped.",
        "PURI_CORRIDORS",
        "Current",
        "LOW",
        "PARTIAL",
        "Institutional ownership classifications documented across 9 major sites",
        "DIRECT_LOCATION_REFERENCE",
        "SRC_SJTA_ACT_1955, SRC_OHREC_LOKANATH",
        "OWN_JAG_001, OWN_GUN_001, OWN_GAP_CADASTRAL_GIS",
        "Revenue & Disaster Management Department Bhulekh GIS parcel integration."
    ],
    [
        "GAP_GIS_001",
        "Official Government Geodetic Benchmark Station Coordinates",
        "Spatial station coordinates are currently captured from secondary geolocation / Google Maps for UI visualization, pending official Survey of India boundary benchmarks.",
        "PURI_SITES",
        "Current",
        "LOW",
        "PARTIAL",
        "9 secondary GPS coordinates documented with explicit SECONDARY_GEOLOCATION label",
        "DIRECT_LOCATION_REFERENCE",
        "SRC_GOOGLE_MAPS_GIS",
        "GIS_JAG_001, GIS_GUN_001, GIS_GAP_OFFICIAL_SURVEY",
        "Survey of India / ORSAC official GIS portal cadastral coordinate verification."
    ]
]

for r in rows1:
    ws1.append(r)

# Sheet 2: GAP_TAXONOMY_RULES
ws2 = wb.create_sheet(title="GAP_TAXONOMY_RULES")
headers2 = ["concept", "allowed_values", "definition", "enforcement_rule"]
rows2 = [
    ["target_status", "AVAILABLE", "Exact target metric is directly supported by verified empirical observations.", "Must have verified primary observation matching exact required geography and definition."],
    ["target_status", "PARTIAL", "Target metric is not directly observed, but related valid evidence, proxies, or broader scopes exist.", "Evidence exists but does NOT close the target gap; cannot be treated as complete."],
    ["target_status", "UNRESOLVED", "No valid empirical observation, proxy, or related evidence exists in package.", "Must be displayed with explicit DATA_GAP status in UI."],
    ["representation_type", "DIRECT", "Primary verified observation measured directly for this indicator.", "Cannot be labelled ESTIMATED or PROXY."],
    ["representation_type", "DERIVED", "Mathematically computed from verified primary observations using published formula.", "Must state formula and parent observation IDs."],
    ["representation_type", "ESTIMATED", "Modelled value using scientific methodology.", "Explicitly prohibited from closing primary empirical data gaps."],
    ["representation_type", "PROXY", "Related empirical measurement from adjacent domain or geographic level used as proxy.", "Must be clearly labelled PROXY with disclaimer."],
    ["representation_type", "CONTEXT_ONLY", "Broader regional, national, or historical baseline providing analytical context.", "Strictly prohibited from being allocated down to destination."],
    ["representation_type", "STRUCTURAL_NOT_APPLICABLE", "Domain area where monitoring infrastructure does not exist at destination.", "Zero synthetic data allowed; transparently disclosed as infrastructure gap."]
]
ws2.append(headers2)
for r in rows2:
    ws2.append(r)

# Sheet 3: METADATA
ws3 = wb.create_sheet(title="METADATA")
ws3.append(["field", "value"])
ws3.append(["destination", "PURI, Odisha"])
ws3.append(["destination_id", "103 (Proposed) / PURI"])
ws3.append(["package_version", "v1.0-contract"])
ws3.append(["total_master_gaps", len(rows1)])
ws3.append(["silently_closed_gaps", 0])
ws3.append(["fabricated_values", 0])
ws3.append(["creation_date", "2026-08-25"])

# Styling
header_fill = PatternFill(start_color="1A381E", end_color="1A381E", fill_type="solid")
header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")

for ws in [ws1, ws2, ws3]:
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

wb.save(out_path)
print(f"Created PURI_DATA_GAPS.xlsx with {len(rows1)} master gaps at {out_path}")
