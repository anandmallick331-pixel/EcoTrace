"""
Konark Destination Ingestion Module.

Authoritative, reproducible ingestion pipeline for Konark data into EcoTrace.
Directly adheres to the s21_ready data package:
- 01_DESTINATION.xlsx
- 02_LOCATIONS.xlsx
- 03_METRIC_DEFINITIONS.xlsx
- 04_SOURCES.xlsx
- 05_DATASETS.xlsx
- 06_OBSERVATIONS.xlsx
- 07_DASHBOARD_SUMMARY.xlsx

Follows existing EcoTrace models, repositories, and services.
Preserves all Konark records using location_id, granular MetricDefinitions,
and full provenance/evidence links while guaranteeing complete non-interference
with Chilika and Bhubaneswar.
"""

from __future__ import annotations

import logging
import os
import re
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

import openpyxl
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.db.session import SessionLocal
from app.models.destination import Destination, Location
from app.models.enums import (
    ConfidenceLevel,
    DestinationSpecificity,
    EvidenceType,
    MetricDirection,
    ObservationStatus,
)
from app.models.evidence import Evidence
from app.models.metric import MetricDefinition
from app.models.observation import Observation
from app.models.source import Dataset, Source
from app.repositories.dataset import DatasetRepository
from app.repositories.destination import DestinationRepository
from app.repositories.evidence import EvidenceRepository
from app.repositories.metric import MetricDefinitionRepository
from app.repositories.observation import ObservationRepository
from app.repositories.source import SourceRepository

logger = logging.getLogger(__name__)


def get_konark_data_dir() -> Path:
    """Locates the Konark s21_ready package directory."""
    candidates = [
        Path(__file__).resolve().parent.parent.parent / "REGENLEDGER_DATA (1)" / "REGENLEDGER_DATA_KONARK_UPDATED" / "s21_ready",
        Path(__file__).resolve().parent.parent.parent / "REGENLEDGER_DATA" / "REGENLEDGER_DATA_KONARK_UPDATED" / "s21_ready",
        Path(__file__).resolve().parent.parent.parent / "REGENLEDGER_DATA (1)" / "Konark" / "s21_ready",
        Path(__file__).resolve().parent.parent.parent / "REGENLEDGER_DATA" / "Konark" / "s21_ready",
    ]
    for c in candidates:
        if c.exists():
            return c
    return candidates[0]


@dataclass
class KonarkIngestionReport:
    """Summary statistics and diagnostic audit for a Konark ingestion run."""
    files_processed: List[str] = field(default_factory=list)
    records_read: int = 0
    records_inserted: int = 0
    records_updated: int = 0
    records_skipped: int = 0
    records_blocked: int = 0
    evidence_created: int = 0
    validation_errors: List[str] = field(default_factory=list)
    unmapped_metrics: List[str] = field(default_factory=list)
    provenance_gaps: List[str] = field(default_factory=list)
    duplicate_preventions: int = 0
    sources_loaded: int = 0
    datasets_loaded: int = 0
    metrics_loaded: int = 0
    locations_loaded: int = 0

    def to_dict(self) -> Dict[str, Any]:
        return {
            "files_processed": self.files_processed,
            "records_read": self.records_read,
            "records_inserted": self.records_inserted,
            "records_updated": self.records_updated,
            "records_skipped": self.records_skipped,
            "records_blocked": self.records_blocked,
            "evidence_created": self.evidence_created,
            "duplicate_preventions": self.duplicate_preventions,
            "sources_loaded": self.sources_loaded,
            "datasets_loaded": self.datasets_loaded,
            "metrics_loaded": self.metrics_loaded,
            "locations_loaded": self.locations_loaded,
            "unmapped_metrics": self.unmapped_metrics,
            "provenance_gaps": self.provenance_gaps,
            "validation_errors": self.validation_errors,
        }


class KonarkIngestionService:
    """
    Domain service orchestrating the ingestion of Konark datasets into EcoTrace.
    """

    def __init__(self, db: Session, data_dir: Optional[Path] = None) -> None:
        self.db = db
        self.data_dir = data_dir or get_konark_data_dir()
        self.dest_repo = DestinationRepository(db)
        self.source_repo = SourceRepository(db)
        self.dataset_repo = DatasetRepository(db)
        self.metric_repo = MetricDefinitionRepository(db)
        self.obs_repo = ObservationRepository(db)
        self.evidence_repo = EvidenceRepository(db)

        self.report = KonarkIngestionReport()

        # Cache for quick lookup during record processing
        self._destination: Optional[Destination] = None
        self._locations: Dict[str, Location] = {}
        self._sources: Dict[str, Source] = {}
        self._datasets: Dict[str, Dataset] = {}
        self._metrics: Dict[str, MetricDefinition] = {}

    # ── Helper: Read Excel Sheet ─────────────────────────────────────────────
    def _read_excel_sheet(self, filename: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
        file_path = self.data_dir / filename
        if not file_path.exists():
            err_msg = f"Konark file not found: {file_path}"
            self.report.validation_errors.append(err_msg)
            logger.error(err_msg)
            return []

        self.report.files_processed.append(filename)
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        header_idx = -1
        for idx, r in enumerate(rows):
            if r and any(c is not None for c in r):
                header_idx = idx
                break

        if header_idx == -1:
            return []

        header = [str(c).strip() if c is not None else "" for c in rows[header_idx]]
        data_rows = []
        for r in rows[header_idx + 1:]:
            if r and any(c is not None for c in r):
                row_dict = dict(zip(header, r))
                data_rows.append(row_dict)

        return data_rows

    # ── Temporal Parsing Helper ──────────────────────────────────────────────
    @staticmethod
    def parse_temporal_period(year_raw: Any, dataset_code: Optional[str] = None) -> Optional[Tuple[date, date]]:
        """
        Parses year or baseline dataset into period_start and period_end dates.
        """
        if year_raw is not None:
            try:
                y = int(year_raw)
                return date(y, 1, 1), date(y, 12, 31)
            except (ValueError, TypeError):
                pass
            
            y_str = str(year_raw).strip()
            fiscal_match = re.match(r"^(\d{4})-(\d{2,4})$", y_str)
            if fiscal_match:
                start_yr = int(fiscal_match.group(1))
                end_group = fiscal_match.group(2)
                end_yr = int(end_group) if len(end_group) == 4 else int(fiscal_match.group(1)[:2] + end_group)
                return date(start_yr, 4, 1), date(end_yr, 3, 31)

        # Baseline mappings for non-annual or static inventory metrics
        if dataset_code == "DS_UNESCO_HERITAGE":
            return date(1984, 1, 1), date(2026, 12, 31)
        elif dataset_code == "DS_HUD_COMMUNITY":
            return date(2024, 1, 1), date(2024, 12, 31)
        elif dataset_code in ("DS_WATER_OUTLAY", "DS_WASTE_PROG"):
            return date(2024, 1, 1), date(2024, 12, 31)
        elif dataset_code == "DS_DERIVED_KONARK":
            return date(2024, 1, 1), date(2024, 12, 31)

        return date(2024, 1, 1), date(2024, 12, 31)

    # ── Geographic Spatial Resolver ──────────────────────────────────────────
    def resolve_location(self, geo_raw: Any) -> Optional[Location]:
        """
        Maps a geographic scope string to a child Location entity.
        Returns None for destination-wide aggregate observations.
        """
        if not geo_raw:
            return None
        geo = str(geo_raw).strip()

        # Destination-wide aggregates
        if geo in [
            "Konark",
            "Destination (ASI ticket)",
            "Destination (ASI ticket) / Konark tourist centre",
            "Konark / relevant scheme geography",
        ]:
            return None

        # Exact match in locations
        if geo in self._locations:
            return self._locations[geo]

        # Alias / partial matching
        if "Sun Temple" in geo:
            for k, loc in self._locations.items():
                if "Sun Temple" in k:
                    return loc
        if "Museum" in geo or "Archaeological" in geo:
            for k, loc in self._locations.items():
                if "Museum" in k:
                    return loc
        if "Chandrabhaga" in geo:
            for k, loc in self._locations.items():
                if "Chandrabhaga" in k:
                    return loc
        if "NAC" in geo or "town centre" in geo:
            for k, loc in self._locations.items():
                if "NAC" in k or "town" in k:
                    return loc
        if "Balukhand" in geo or "Sanctuary" in geo:
            for k, loc in self._locations.items():
                if "Balukhand" in k or "Sanctuary" in k:
                    return loc

        return None

    # ── 1. Destination & Locations Setup ─────────────────────────────────────
    def setup_destination_and_locations(self) -> Destination:
        """
        Creates or fetches the Konark Destination and its spatial child Locations.
        """
        dest_rows = self._read_excel_sheet("01_DESTINATION.xlsx", "DESTINATION")
        dest_name = "Konark"
        dest_desc = (
            "Konark – UNESCO World Heritage Sun Temple destination, Odisha. "
            "Regenerative tourism data package (REGENLEDGER)."
        )
        country_code = "IND"
        region = "Odisha"

        if dest_rows:
            r = dest_rows[0]
            dest_name = str(r.get("name") or dest_name).strip()
            dest_desc = str(r.get("description") or dest_desc).strip()
            country_code = str(r.get("country_code") or country_code).strip()
            region = str(r.get("region") or region).strip()

        dest = self.dest_repo.get_by_name(dest_name)
        if not dest:
            dest = self.dest_repo.create(
                name=dest_name,
                description=dest_desc,
                country_code=country_code,
                region=region,
            )
            logger.info(f"Created Destination: {dest.name} (id={dest.id})")
        self._destination = dest

        # Read 02_LOCATIONS.xlsx
        loc_rows = self._read_excel_sheet("02_LOCATIONS.xlsx", "LOCATIONS")
        spatial_locations: List[Tuple[str, Optional[float], Optional[float], Optional[str]]] = []

        for r in loc_rows:
            label = str(r.get("label", "")).strip()
            if not label:
                continue
            lat = float(r["latitude"]) if r.get("latitude") is not None and str(r["latitude"]).strip() != "" else None
            lon = float(r["longitude"]) if r.get("longitude") is not None and str(r["longitude"]).strip() != "" else None
            desc = str(r.get("geo_scope_notes", "")).strip() if r.get("geo_scope_notes") else None
            spatial_locations.append((label, lat, lon, desc))

        existing_locs = {loc.label: loc for loc in dest.locations if loc.label}
        for label, lat, lon, desc in spatial_locations:
            if label not in existing_locs:
                loc = Location(
                    destination_id=dest.id,
                    label=label,
                    latitude=lat,
                    longitude=lon,
                    geojson=None,
                )
                self.db.add(loc)
                existing_locs[label] = loc
                self.report.locations_loaded += 1

        self.db.commit()
        self.db.refresh(dest)
        self._locations = {loc.label: loc for loc in dest.locations if loc.label}
        return dest

    # ── 2. Source Register & Datasets Ingestion ──────────────────────────────
    def load_sources_and_datasets(self) -> None:
        """
        Loads the sources and datasets from 04_SOURCES.xlsx and 05_DATASETS.xlsx.
        """
        src_rows = self._read_excel_sheet("04_SOURCES.xlsx", "SOURCES")
        
        # Deduplicate sources by source_code, keeping best metadata
        deduped_sources: Dict[str, Dict[str, Any]] = {}
        for r in src_rows:
            s_code = str(r.get("source_code", "")).strip()
            if not s_code:
                continue
            if s_code not in deduped_sources:
                deduped_sources[s_code] = r
            else:
                if not deduped_sources[s_code].get("official_url") and r.get("official_url"):
                    deduped_sources[s_code]["official_url"] = r["official_url"]
                if r.get("notes"):
                    existing_notes = deduped_sources[s_code].get("notes") or ""
                    if r["notes"] not in existing_notes:
                        deduped_sources[s_code]["notes"] = f"{existing_notes} | {r['notes']}".strip(" |")

        for s_code, r in deduped_sources.items():
            auth = r.get("authority")
            doc_title = r.get("document_title")
            notes = r.get("notes")
            url = r.get("official_url")

            full_desc = f"{doc_title}. {notes}" if (doc_title and notes) else (doc_title or notes or s_code)

            source = self.source_repo.get_by_name(s_code)
            if not source:
                source = self.source_repo.create(
                    name=s_code,
                    organisation=str(auth) if auth else None,
                    description=str(full_desc) if full_desc else None,
                    url=str(url) if url else None,
                )
                self.report.sources_loaded += 1
            self._sources[s_code] = source

        ds_rows = self._read_excel_sheet("05_DATASETS.xlsx", "DATASETS")
        for r in ds_rows:
            ds_code = str(r.get("dataset_code", "")).strip()
            s_code = str(r.get("source_code", "")).strip()
            name = str(r.get("name", "")).strip()
            version = str(r.get("version", "")).strip() if r.get("version") else None
            pub_date_raw = r.get("publication_date")
            desc = str(r.get("description", "")).strip() if r.get("description") else None

            pub_date: Optional[date] = None
            if pub_date_raw:
                try:
                    if isinstance(pub_date_raw, date):
                        pub_date = pub_date_raw
                    else:
                        pub_date = date.fromisoformat(str(pub_date_raw).strip())
                except (ValueError, TypeError):
                    pub_date = None

            src = self._sources.get(s_code)
            if not src:
                self.report.validation_errors.append(f"Source {s_code} not found for dataset {ds_code}")
                continue

            stmt = select(Dataset).where(Dataset.name == name, Dataset.source_id == src.id)
            ds = self.db.scalars(stmt).first()
            if not ds:
                ds = self.dataset_repo.create(
                    name=name,
                    source_id=src.id,
                    version=version,
                    publication_date=pub_date,
                    description=f"{ds_code}: {desc}" if desc else ds_code,
                )
                self.report.datasets_loaded += 1
            self._datasets[ds_code] = ds
            self._datasets[name] = ds

    # ── 3. Metric Definitions Synchronization ───────────────────────────────
    def load_metric_definitions(self) -> None:
        """
        Loads the 31 metric definitions from 03_METRIC_DEFINITIONS.xlsx.
        """
        met_rows = self._read_excel_sheet("03_METRIC_DEFINITIONS.xlsx", "METRIC_DEFINITIONS")
        direction_map = {
            "HIGHER_IS_BETTER": MetricDirection.HIGHER_IS_BETTER,
            "LOWER_IS_BETTER": MetricDirection.LOWER_IS_BETTER,
            "NEUTRAL": MetricDirection.NEUTRAL,
        }

        for r in met_rows:
            code = str(r.get("code", "")).strip()
            if not code:
                continue
            version = str(r.get("version", "1.0")).strip() or "1.0"
            name = str(r.get("name", "")).strip()
            category = str(r.get("category", "")).strip()
            unit = str(r.get("unit", "")).strip()
            raw_dir = str(r.get("direction", "NEUTRAL")).strip().upper()
            direction = direction_map.get(raw_dir, MetricDirection.NEUTRAL)
            description = str(r.get("description", "")).strip() if r.get("description") else None

            m = self.metric_repo.get_by_code_version(code, version)
            if not m:
                m = self.metric_repo.create(
                    code=code,
                    version=version,
                    name=name,
                    category=category,
                    unit=unit,
                    direction=direction,
                    description=description,
                )
                self.report.metrics_loaded += 1
            self._metrics[code] = m

    # ── 4. Observation & Evidence Ingestion ──────────────────────────────────
    def ingest_observations(self) -> None:
        """
        Ingests observations and provenance evidence from 06_OBSERVATIONS.xlsx.
        """
        obs_rows = self._read_excel_sheet("06_OBSERVATIONS.xlsx", "OBSERVATIONS")
        if not obs_rows:
            self.report.validation_errors.append("No observation records found in 06_OBSERVATIONS.xlsx")
            return

        for rec in obs_rows:
            self.report.records_read += 1
            metric_code = str(rec.get("metric_code", "")).strip()
            year_raw = rec.get("year")
            raw_val = rec.get("value")
            raw_unit = str(rec.get("unit", "")).strip()
            raw_status = str(rec.get("status", "")).strip().upper()
            raw_conf = str(rec.get("confidence", "")).strip().upper()
            raw_val_type = str(rec.get("value_type", "")).strip().upper()
            calc_formula = str(rec.get("calculation_formula", "")).strip() if rec.get("calculation_formula") else None
            input_ids = str(rec.get("input_record_ids", "")).strip() if rec.get("input_record_ids") else None
            source_code_raw = str(rec.get("source_code", "")).strip() if rec.get("source_code") else None
            dataset_code = str(rec.get("dataset_code", "")).strip() if rec.get("dataset_code") else None
            raw_notes = str(rec.get("notes", "")).strip() if rec.get("notes") else ""
            geo_scope = str(rec.get("geographic_scope", "")).strip()

            is_data_gap = (raw_status == "DATA_GAP" or raw_val_type == "DATA_GAP")

            # ── Resolve Metric Definition ──
            metric_def = self._metrics.get(metric_code)
            if not metric_def:
                err_msg = f"Metric code '{metric_code}' not defined in database"
                self.report.validation_errors.append(err_msg)
                self.report.records_skipped += 1
                continue

            # ── Resolve Dataset ──
            dataset: Optional[Dataset] = None
            if dataset_code:
                dataset = self._datasets.get(dataset_code)
                if not dataset:
                    err_msg = f"Dataset code '{dataset_code}' not defined in database"
                    self.report.validation_errors.append(err_msg)
                    self.report.records_skipped += 1
                    continue
            elif not is_data_gap:
                err_msg = f"Non-gap observation missing dataset_code for metric {metric_code}"
                self.report.validation_errors.append(err_msg)
                self.report.records_skipped += 1
                continue

            dataset_id = dataset.id if dataset else None

            # ── Resolve Spatial Location ──
            loc_obj = self.resolve_location(geo_scope)
            loc_id = loc_obj.id if loc_obj else None

            # ── Parse Temporal Period ──
            period = self.parse_temporal_period(year_raw, dataset_code)
            if not period:
                self.report.records_skipped += 1
                self.report.validation_errors.append(f"Unparseable date period '{year_raw}' for metric {metric_code}")
                continue

            period_start, period_end = period

            # ── Parse Numeric Values (preserve non-numeric qualitative strings in notes) ──
            orig_value: Optional[float] = None
            norm_value: Optional[float] = None
            if raw_val is not None:
                try:
                    if isinstance(raw_val, (int, float)):
                        orig_value = float(raw_val)
                    else:
                        orig_value = float(str(raw_val).strip())
                    norm_value = orig_value
                except (ValueError, TypeError):
                    orig_value = None
                    norm_value = None

            # ── Status & Confidence Mapping ──
            obs_status = ObservationStatus.VERIFIED
            if is_data_gap or "RAW" in raw_status or "FLAGGED" in raw_status:
                obs_status = ObservationStatus.RAW

            confidence = ConfidenceLevel.HIGH
            if "MEDIUM" in raw_conf:
                confidence = ConfidenceLevel.MEDIUM
            elif "LOW" in raw_conf:
                confidence = ConfidenceLevel.LOW
            elif "UNKNOWN" in raw_conf or "N/A" in raw_conf or is_data_gap:
                confidence = ConfidenceLevel.UNKNOWN

            dest_spec = DestinationSpecificity.DIRECT
            if "STATEWIDE" in geo_scope.upper() or "ODISHA" in geo_scope.upper() and "KONARK" not in geo_scope.upper():
                dest_spec = DestinationSpecificity.REGIONAL
            elif raw_val_type in ("DERIVED", "ESTIMATED", "PROXY", "ILLUSTRATIVE"):
                dest_spec = DestinationSpecificity.MODELLED

            # ── Construct Structured Notes & Trace ──
            notes_payload = (
                f"metric_code: {metric_code} | "
                f"year: {year_raw} | "
                f"raw_value: {raw_val} | "
                f"raw_unit: {raw_unit} | "
                f"status: {raw_status} | "
                f"confidence: {raw_conf} | "
                f"value_type: {raw_val_type} | "
                f"calculation_formula: {calc_formula} | "
                f"input_record_ids: {input_ids} | "
                f"source_code: {source_code_raw} | "
                f"dataset_code: {dataset_code} | "
                f"geographic_scope: {geo_scope} | "
                f"notes: {raw_notes}"
            )

            methodology_payload = (
                f"Type: {raw_val_type} | Formula: {calc_formula} | Inputs: {input_ids}"
                if (calc_formula or input_ids)
                else (
                    f"Type: {raw_val_type} | Absence of verified empirical evidence (DATA_GAP)"
                    if is_data_gap
                    else f"Type: {raw_val_type} | Direct observation from {dataset_code}"
                )
            )

            # ── Natural Key Collision & Idempotency Check ──
            # uq_observation_natural_key: (destination_id, location_id, metric_definition_id, dataset_id, period_start, period_end)
            stmt = select(Observation).where(
                Observation.destination_id == self._destination.id,
                Observation.location_id == loc_id,
                Observation.metric_definition_id == metric_def.id,
                Observation.dataset_id == dataset_id,
                Observation.period_start == period_start,
                Observation.period_end == period_end,
            )
            existing_obs = self.db.scalars(stmt).first()

            if existing_obs:
                self.report.duplicate_preventions += 1
                if existing_obs.original_value is None and orig_value is not None:
                    existing_obs.original_value = orig_value
                    existing_obs.normalized_value = norm_value
                    existing_obs.status = obs_status
                    existing_obs.confidence = confidence
                    existing_obs.notes = notes_payload
                    self.db.add(existing_obs)
                    self.db.commit()
                    self.report.records_updated += 1
                obs = existing_obs
            else:
                # ── Create Observation ──
                obs = Observation(
                    destination_id=self._destination.id,
                    location_id=loc_id,
                    metric_definition_id=metric_def.id,
                    dataset_id=dataset_id,
                    period_start=period_start,
                    period_end=period_end,
                    original_value=orig_value,
                    normalized_value=norm_value,
                    status=obs_status,
                    confidence=confidence,
                    destination_specificity=dest_spec,
                    methodology=methodology_payload,
                    assumptions=f"Status: {raw_status} | ValueType: {raw_val_type}",
                    notes=notes_payload,
                )
                self.db.add(obs)
                self.db.commit()
                self.db.refresh(obs)
                self.report.records_inserted += 1

            # ── Evidence Creation (Supports single and multi-source codes) ──
            if source_code_raw and not is_data_gap:
                source_codes = [s.strip() for s in source_code_raw.split(";") if s.strip()]
                for s_code in source_codes:
                    source_obj = self._sources.get(s_code)
                    if source_obj:
                        ev_stmt = select(Evidence).where(
                            Evidence.observation_id == obs.id,
                            Evidence.source_id == source_obj.id,
                        )
                        existing_ev = self.db.scalars(ev_stmt).first()
                        if not existing_ev:
                            ev = Evidence(
                                observation_id=obs.id,
                                source_id=source_obj.id,
                                evidence_type=EvidenceType.DOCUMENT,
                                reference_url=source_obj.url,
                                raw_excerpt=raw_notes[:500] if raw_notes else f"Direct observation from {s_code}",
                                notes=f"Source: {s_code} | Dataset: {dataset_code} | Formula: {calc_formula or 'N/A'}",
                            )
                            self.db.add(ev)
                            self.report.evidence_created += 1

        self.db.commit()

    # ── 5. Safe Teardown Helper ──────────────────────────────────────────────
    def clear_konark_observations(self) -> int:
        """
        Safely removes existing Konark observations and evidence created by prior runs.
        Guarantees that Chilika (id=44) and Bhubaneswar (id=100) are never touched.
        """
        dest = self.dest_repo.get_by_name("Konark")
        if not dest:
            return 0

        # Find all observation IDs for Konark only
        obs_stmt = select(Observation.id).where(Observation.destination_id == dest.id)
        konark_obs_ids = list(self.db.scalars(obs_stmt).all())

        if konark_obs_ids:
            # Delete evidence items for Konark observations
            del_ev_stmt = delete(Evidence).where(Evidence.observation_id.in_(konark_obs_ids))
            self.db.execute(del_ev_stmt)

            # Delete Konark observations
            del_obs_stmt = delete(Observation).where(Observation.destination_id == dest.id)
            self.db.execute(del_obs_stmt)
            self.db.commit()
            logger.info(f"Cleared {len(konark_obs_ids)} existing Konark observations and associated evidence.")

        return len(konark_obs_ids)

    # ── Master Ingestion Pipeline ────────────────────────────────────────────
    def run(self, clean_existing_konark: bool = False) -> KonarkIngestionReport:
        """
        Executes the full Konark pilot ingestion pipeline in sequential, idempotent steps.
        """
        logger.info("Starting Konark Ingestion Pipeline...")
        self.report = KonarkIngestionReport()

        try:
            if clean_existing_konark:
                self.clear_konark_observations()

            self.setup_destination_and_locations()
            self.load_sources_and_datasets()
            self.load_metric_definitions()
            self.ingest_observations()

            logger.info(
                f"Konark Ingestion Completed. Inserted: {self.report.records_inserted}, "
                f"Sources: {self.report.sources_loaded}, Datasets: {self.report.datasets_loaded}, "
                f"Metrics: {self.report.metrics_loaded}, Locations: {self.report.locations_loaded}"
            )
        except Exception as e:
            self.db.rollback()
            err_msg = f"Fatal error during Konark ingestion: {str(e)}"
            logger.error(err_msg, exc_info=True)
            self.report.validation_errors.append(err_msg)
            raise

        return self.report


def run_konark_ingestion(clean_existing: bool = False) -> KonarkIngestionReport:
    """CLI/Helper entry point to execute the Konark ingestion pipeline."""
    db = SessionLocal()
    try:
        service = KonarkIngestionService(db)
        report = service.run(clean_existing_konark=clean_existing)
        return report
    finally:
        db.close()


if __name__ == "__main__":
    rep = run_konark_ingestion()
    print("=" * 60)
    print("KONARK INGESTION EXECUTION REPORT")
    print("=" * 60)
    for k, v in rep.to_dict().items():
        print(f"  {k}: {v}")
