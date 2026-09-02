"""
Chilika Pilot Ingestion Module.

Authoritative, reproducible ingestion pipeline for Chilika Lake data into EcoTrace.
Directly adheres to:
- pipeline/CHILIKA_INGESTION_MAPPING.xlsx
- metadata/SOURCE_REGISTER.xlsx
- framework/METRIC_DICTIONARY (1).xlsx
- framework/DIRECTION_SCORING_FRAMEWORK_CORRECTED (2).xlsx
- quality_reports/*.xlsx

Follows existing EcoTrace models, repositories, and services.
Preserves all 533 valid records using location_id and granular MetricDefinitions.
"""

from __future__ import annotations

import logging
import os
import re
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

import openpyxl
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.db.session import SessionLocal
from app.models.destination import Destination, Location
from app.models.enums import (
    ConfidenceLevel,
    DestinationSpecificity,
    EvidenceType,
    MetricDirection,
    ObservationStatus,
)
from app.models.evidence import Evidence
from app.models.metric import MetricDefinition
from app.models.observation import Observation
from app.models.source import Dataset, Source
from app.repositories.dataset import DatasetRepository
from app.repositories.destination import DestinationRepository
from app.repositories.evidence import EvidenceRepository
from app.repositories.metric import MetricDefinitionRepository
from app.repositories.observation import ObservationRepository
from app.repositories.source import SourceRepository

logger = logging.getLogger(__name__)

def get_chilika_data_dir() -> Path:
    candidates = [
        Path(__file__).resolve().parent.parent.parent / "REGENLEDGER_DATA (1)" / "REGENLEDGER_DATA_CHILIKA_UPDATED",
        Path(__file__).resolve().parent.parent.parent / "REGENLEDGER_DATA" / "REGENLEDGER_DATA_CHILIKA_UPDATED",
        Path(__file__).resolve().parent.parent.parent / "REGENLEDGER_DATA (1)" / "REGENLEDGER_DATA",
        Path(__file__).resolve().parent.parent.parent / "REGENLEDGER_DATA",
    ]
    for c in candidates:
        if c.exists():
            return c
    return candidates[0]

DATA_BASE_DIR = get_chilika_data_dir()


@dataclass
class IngestionReport:
    """Summary statistics and diagnostic audit for a Chilika ingestion run."""
    files_processed: List[str] = field(default_factory=list)
    records_read: int = 0
    records_inserted: int = 0
    records_updated: int = 0
    records_skipped: int = 0
    records_blocked: int = 0
    evidence_created: int = 0
    validation_errors: List[str] = field(default_factory=list)
    unmapped_metrics: List[str] = field(default_factory=list)
    provenance_gaps: List[str] = field(default_factory=list)
    duplicate_preventions: int = 0
    sources_loaded: int = 0
    datasets_loaded: int = 0
    metrics_loaded: int = 0
    locations_loaded: int = 0

    def to_dict(self) -> Dict[str, Any]:
        return {
            "files_processed": self.files_processed,
            "records_read": self.records_read,
            "records_inserted": self.records_inserted,
            "records_updated": self.records_updated,
            "records_skipped": self.records_skipped,
            "records_blocked": self.records_blocked,
            "evidence_created": self.evidence_created,
            "duplicate_preventions": self.duplicate_preventions,
            "sources_loaded": self.sources_loaded,
            "datasets_loaded": self.datasets_loaded,
            "metrics_loaded": self.metrics_loaded,
            "locations_loaded": self.locations_loaded,
            "unmapped_metrics": self.unmapped_metrics,
            "provenance_gaps": self.provenance_gaps,
            "validation_errors": self.validation_errors,
        }


class ChilikaIngestionService:
    """
    Domain service orchestrating the ingestion of Chilika pilot datasets.
    """

    def __init__(self, db: Session) -> None:
        self.db = db
        self.dest_repo = DestinationRepository(db)
        self.source_repo = SourceRepository(db)
        self.dataset_repo = DatasetRepository(db)
        self.metric_repo = MetricDefinitionRepository(db)
        self.obs_repo = ObservationRepository(db)
        self.evidence_repo = EvidenceRepository(db)

        self.report = IngestionReport()

        # Cache for quick lookup during record processing
        self._destination: Optional[Destination] = None
        self._locations: Dict[str, Location] = {}
        self._sources: Dict[str, Source] = {}
        self._datasets: Dict[str, Dataset] = {}
        self._metrics: Dict[str, MetricDefinition] = {}
        self._blocked_obs_ids: Set[str] = {
            "BIO-IND-SEAGRASS-6",  # BLK-REC-001
            "BIO-IND-TOTAL-383",   # BLK-MET-002 (needs_metric_definition)
            "FIS-COMM-001",        # BLK-MET-001 (fisher_population)
            "FIS-IND-010",         # BLK-MET-002 (needs_metric_definition)
            "FIS-SPP-003",         # BLK-MET-002 (needs_metric_definition)
            "FIS-SPP-004",         # BLK-MET-002 (needs_metric_definition)
        }
        self._blocked_metrics: Set[str] = {
            "fisher_population",       # BLK-MET-001 / AMB-MET-001 (name implies count, content is income)
            "needs_metric_definition", # BLK-MET-002 / AMB-MET-002 (heterogeneous placeholder)
            "seagrass_richness",       # BLK-REC-001 / AMB-REC-001 (needs_review)
        }

    # ── Temporal Parsing Helper ──────────────────────────────────────────────
    @staticmethod
    def parse_temporal_period(year_raw: Any) -> Optional[Tuple[date, date]]:
        """
        Parses year/period string according to DATE_PERIOD_MAPPING in CHILIKA_INGESTION_MAPPING.xlsx.
        Returns (period_start, period_end) or None if unparseable/blocked.
        """
        if year_raw is None:
            return None
        
        y_str = str(year_raw).strip()

        # Fiscal / reporting year patterns (e.g. 2023-24, 2021-22)
        fiscal_match = re.match(r"^(\d{4})-(\d{2})(?:\s*\(.*\))?$", y_str)
        if fiscal_match:
            start_yr = int(fiscal_match.group(1))
            end_yr = int(fiscal_match.group(1)[:2] + fiscal_match.group(2))
            return date(start_yr, 4, 1), date(end_yr, 3, 31)

        # Standard 4-digit calendar year (e.g. 2024, 2023, 2020)
        cal_match = re.match(r"^(\d{4})$", y_str)
        if cal_match:
            yr = int(cal_match.group(1))
            return date(yr, 1, 1), date(yr, 12, 31)

        # Multi-year spans & historical ranges
        multi_patterns: Dict[str, Tuple[date, date]] = {
            "2021-2022": (date(2021, 1, 1), date(2022, 12, 31)),
            "2021-2022 (average)": (date(2021, 1, 1), date(2022, 12, 31)),
            "2021-2022 (avg annual)": (date(2021, 1, 1), date(2022, 12, 31)),
            "2023-2024": (date(2023, 1, 1), date(2024, 12, 31)),
            "2023-2024 (average)": (date(2023, 4, 1), date(2024, 3, 31)),
            "2019-2020 (avg annual, comparison baseline)": (date(2019, 4, 1), date(2020, 3, 31)),
            "2011-12 & 2012-13": (date(2011, 4, 1), date(2013, 3, 31)),
            "2011-12 to 2012-13": (date(2011, 4, 1), date(2013, 3, 31)),
            "2010-11 to 2012-13": (date(2010, 4, 1), date(2013, 3, 31)),
            "2000-08": (date(2000, 1, 1), date(2008, 12, 31)),
            "2011-2013": (date(2011, 1, 1), date(2013, 12, 31)),
            "2012-2020 (updated)": (date(2012, 1, 1), date(2020, 12, 31)),
            "up to 2012-13": (date(2010, 4, 1), date(2013, 3, 31)),
            "up to 2016 (pub 2018)": (date(1916, 1, 1), date(2016, 12, 31)),
            "up to 2016 (published 2018/2019)": (date(1916, 1, 1), date(2016, 12, 31)),
            "1916-2014 (Mohanty 2015)": (date(1916, 1, 1), date(2014, 12, 31)),
            "Mohapatra et al. 2015": (date(1916, 1, 1), date(2015, 12, 31)),
            "earlier RPRC": (date(1980, 1, 1), date(2012, 12, 31)),
            "ongoing": (date(1987, 1, 1), date(2024, 12, 31)),
            "Current": (date(2024, 1, 1), date(2024, 12, 31)),
            "~2023 (guideline date)": (date(2023, 1, 1), date(2023, 12, 31)),
            "threshold": (date(2005, 1, 1), date(2005, 12, 31)),
            "threshold (CIFRI 2005)": (date(2005, 1, 1), date(2005, 12, 31)),
            "—": (date(2005, 1, 1), date(2005, 12, 31)),
            "–": (date(2005, 1, 1), date(2005, 12, 31)),
            "-": (date(2005, 1, 1), date(2005, 12, 31)),
            "": (date(2005, 1, 1), date(2005, 12, 31)),
        }

        for pattern, span in multi_patterns.items():
            if y_str.lower() == pattern.lower() or y_str.startswith(pattern):
                return span

        return None

    # ── Geographic Spatial Resolver ──────────────────────────────────────────
    def resolve_location(self, geo_raw: Any) -> Optional[Location]:
        """
        Maps a raw geography string from Chilika datasets to a child Location entity.
        Returns None for destination-wide lagoon observations.
        """
        if not geo_raw:
            return None
        geo = str(geo_raw).strip()
        if geo in [
            "Chilika Lagoon",
            "Chilika Lake",
            "Chilika (brackishwater)",
            "Chilika",
            "Chilika Circle",
            "Chilika - TOTAL CHILIKA",
        ]:
            return None

        # Check for monitoring station format: 'Chilika Station X (lat, lon)' or 'Nalabana NBX (lat, lon)'
        m = re.search(r"(.*?)\s*\(", geo)
        if m:
            st_name = m.group(1).strip()
            if st_name in self._locations:
                return self._locations[st_name]

        # Multi-hub composite strings (e.g. Satapada and Barkul area)
        if "Satapada" in geo and "Barkul" in geo:
            return None

        # Exact / Substring matches
        if "Panthanivas Barkul" in geo and "Panthanivas Barkul" in self._locations:
            return self._locations["Panthanivas Barkul"]
        if "Panthanivas Rambha" in geo and "Panthanivas Rambha" in self._locations:
            return self._locations["Panthanivas Rambha"]
        if "Barkul" in geo and "Barkul" in self._locations:
            return self._locations["Barkul"]
        if "Satapada" in geo and "Satapada" in self._locations:
            return self._locations["Satapada"]
        if "Rambha" in geo and "Rambha" in self._locations:
            return self._locations["Rambha"]
        if "Nalabana" in geo and "Nalabana" in self._locations:
            return self._locations["Nalabana"]
        if "Khurda" in geo and "Chilika - Khurda" in self._locations:
            return self._locations["Chilika - Khurda"]
        if "Puri" in geo and "Chilika - Puri" in self._locations:
            return self._locations["Chilika - Puri"]
        if "Ganjam" in geo and "Chilika - Ganjam" in self._locations:
            return self._locations["Chilika - Ganjam"]
        if "Southern Zone" in geo and "Southern Zone" in self._locations:
            return self._locations["Southern Zone"]
        if "Central Zone" in geo and "Central Zone" in self._locations:
            return self._locations["Central Zone"]
        if "Northern Zone" in geo and "Northern Zone" in self._locations:
            return self._locations["Northern Zone"]
        if "Outer Channel" in geo and "Outer Channel Zone" in self._locations:
            return self._locations["Outer Channel Zone"]

        return None

    # ── Metric Resolution Helper ─────────────────────────────────────────────
    def resolve_metric_code(self, rec: Dict[str, Any], filename: str) -> str:
        """
        Resolves a record's indicator and coarse metric to a granular MetricDefinition code.
        """
        ind = str(rec.get("indicator", "")).strip()
        m_key = str(rec.get("regenledger_metric", "")).strip()

        if filename == "chilika_water_processed.xlsx":
            if "Fecal Coliform" in ind: return "water_fecal_coliform"
            if "Total Coliform" in ind: return "water_total_coliform"
            if "pH" in ind: return "water_ph"
            if "Dissolved Oxygen" in ind: return "water_dissolved_oxygen"
            if "BOD" in ind: return "water_bod"
            if "Temperature" in ind: return "water_temperature"
            return "water_quality_parameter"

        if filename == "chilika_fisheries_processed.xlsx":
            if "Total catch" in ind or "Average annual total landings" in ind: return "fish_landings_total"
            if "landings value" in ind: return "fish_landings_value"
            if ("Fish" in ind or "finfish" in ind or "fish" in ind) and ("composition" in ind or "%" in ind): return "fish_composition_finfish"
            if ("Prawn" in ind or "prawn" in ind) and ("composition" in ind or "%" in ind): return "fish_composition_prawn"
            if ("Crab" in ind or "crab" in ind) and ("composition" in ind or "%" in ind): return "fish_composition_crab"
            if "Fish production" in ind or "Production (MT)" in ind: return "fish_production_annual"
            if "health grade" in ind or "index grade" in ind: return "fisheries_health_grade"
            if "commercial species" in ind: return "commercial_species_diversity"
            if "Mugil cephalus" in ind: return "sustainable_size_mullet"
            if "Penaeus monodon" in ind: return "sustainable_size_prawn"
            if "Scylla serrata" in ind: return "sustainable_size_crab"
            if "MSY" in ind or "Maximum Sustainable Yield" in ind: return "biodiversity_msy_threshold"
            if "Total finfish species" in ind: return "species_richness_finfish"
            if "Total shellfish species" in ind: return "species_richness_shellfish"
            if "Near Threatened" in ind: return "near_threatened_finfish_species"
            if "Threatened finfish" in ind: return "threatened_finfish_species"
            if "Not Evaluated" in ind: return "species_richness_not_evaluated"
            return m_key

        if filename == "chilika_biodiversity_processed.xlsx":
            if "MSY" in ind or "Maximum Sustainable Yield" in ind: return "biodiversity_msy_threshold"
            if "commercial species" in ind: return "commercial_species_diversity"
            if "Mugil cephalus" in ind: return "sustainable_size_mullet"
            if "Penaeus monodon" in ind: return "sustainable_size_prawn"
            if "Scylla serrata" in ind: return "sustainable_size_crab"
            if "Report Card grade" in ind or "health grade" in ind: return "ecosystem_health_grade"
            if "Avifauna" in ind: return "bird_species_richness_study"
            if "Least Concern" in ind: return "birds_least_concern"
            if "Near Threatened" in ind: return "birds_near_threatened"
            if "Vulnerable" in ind: return "birds_vulnerable"
            if "Endangered" in ind: return "birds_endangered"
            if "Fishing cat" in ind: return "fishing_cat_population"
            if "bird population" in ind: return "total_bird_census_count"
            if "Flowering plant" in ind or "flowering plant" in ind: return "floral_species_richness_angiosperm"
            if "Angiosperm species" in ind: return "floral_species_richness_angiosperm_rprc"
            if "Phytoplankton" in ind: return "phytoplankton_species_richness"
            if "Sanctuary core" in ind or "Sanctuary area" in ind or "Nalabana Island" in ind: return "nalabana_sanctuary_area"
            if "Finfish species" in ind: return "species_richness_finfish"
            if "Shellfish species" in ind: return "species_richness_shellfish"
            return m_key

        if filename == "chilika_community_processed.xlsx":
            if "villages" in ind and "Fishermen" in ind: return "community_fisher_villages"
            if "households" in ind: return "community_fisher_households"
            if "Total fishermen population" in ind or "fishermen population" in ind: return "community_fisher_population"
            if "solely dependent" in ind: return "community_fishers_solely_dependent"
            if "per-capita income" in ind: return "community_fisher_income_per_capita"
            if "PFCSs provided" in ind: return "community_pfcs_soft_loans"
            if "soft loan amount" in ind: return "community_loan_disbursement"
            if "state-wide FISHFED" in ind or "FISHFED" in ind: return "community_fishfed_pfcs_count"
            if "Boxes (IFB) total" in ind: return "community_ifb_total_boxes_supplied"
            if "Boxes (IFB) supplied" in ind: return "community_ifb_boxes_supplied"
            if "fishers benefited by IFB" in ind: return "community_ifb_fishers_benefited"
            if "price realization" in ind: return "community_ifb_price_realization"
            if "MPEDA subsidy" in ind: return "community_ifb_mpeda_subsidy"
            if "CDA subsidy" in ind: return "community_ifb_cda_subsidy"
            if "Traditional fishers trained" in ind or "fishers trained" in ind: return "community_fishers_trained"
            if "Boatmen trained" in ind: return "trained_boatmen_count"
            if "Training camps" in ind: return "community_training_camps"
            if "Fishing villages covered" in ind: return "community_training_villages"
            if "PFCSs covered by training" in ind: return "community_training_pfcs_covered"
            return m_key

        if filename == "chilika_tourism_processed.xlsx":
            if "boatmen" in ind.lower(): return "trained_boatmen_count"
            if "day cruise" in ind.lower() and ("number" in ind.lower() or "vessel" in ind.lower() or "count" in ind.lower()): return "cruise_vessels_count"
            if "seating capacity" in ind.lower() or "capacity of day cruise" in ind.lower(): return "cruise_seating_capacity"
            if "houseboat" in ind.lower(): return "houseboat_vessels_count"
            if "ferry" in ind.lower(): return "ferry_vessels_count"
            if "identified tourist centre" in ind.lower(): return "identified_tourist_centres_count"
            if "hotel rooms" in ind.lower() or "rooms capacity" in ind.lower(): return "hotel_capacity_rooms"
            if "hotel beds" in ind.lower() or "beds capacity" in ind.lower(): return "hotel_capacity_beds"
            if "hotels/facilities" in ind.lower() or "number of hotels" in ind.lower(): return "hotel_facilities_count"
            if "occupancy rate" in ind.lower() or "percentage of hotel occupancy" in ind.lower(): return "hotel_occupancy_rate"
            if "domestic" in ind.lower(): return "tourist_footfall_domestic"
            if "foreign" in ind.lower(): return "tourist_footfall_foreign"
            if "total tourist footfall" in ind.lower(): return "tourist_footfall_total"
            return m_key

        return m_key

    # ── 1. Destination & Locations Setup ─────────────────────────────────────
    def setup_destination_and_locations(self) -> Destination:
        """
        Creates or fetches the Chilika Destination and its spatial child Locations (52 total).
        """
        dest = self.dest_repo.get_by_name("Chilika")
        if not dest:
            dest = self.dest_repo.create(
                name="Chilika",
                description="Chilika Lagoon - Ramsar wetland of international importance (Site No. 229) and EcoTrace pilot destination.",
                country_code="IND",
                region="Odisha",
            )
            logger.info(f"Created Destination: {dest.name} (id={dest.id})")
        self._destination = dest

        # Core spatial locations from DESTINATION_GEOGRAPHY_MAPPING
        spatial_locations = [
            ("Satapada", 19.6756, 85.4412, "Outer channel tourism hub and dolphin watching base"),
            ("Barkul", 19.6897, 85.1952, "Central lagoon tourism hub and water sports base"),
            ("Rambha", 19.5167, 85.1000, "Southern lagoon tourism hub and island excursion base"),
            ("Nalabana", 19.6917, 85.2917, "Nalabana Bird Sanctuary island (protected area)"),
            ("Panthanivas Barkul", 19.6897, 85.1952, "OTDC accommodation facility at Barkul"),
            ("Panthanivas Rambha", 19.5167, 85.1000, "OTDC accommodation facility at Rambha"),
            ("Chilika - Khurda", 19.8000, 85.3000, "Khurda district administrative sector of Chilika"),
            ("Chilika - Puri", 19.7500, 85.5000, "Puri district administrative sector of Chilika"),
            ("Chilika - Ganjam", 19.5000, 85.1000, "Ganjam district administrative sector of Chilika"),
            ("Southern Zone", 19.5200, 85.1500, "CDA Southern ecological zone"),
            ("Central Zone", 19.6800, 85.2800, "CDA Central ecological zone"),
            ("Northern Zone", 19.8000, 85.3500, "CDA Northern ecological zone"),
            ("Outer Channel Zone", 19.6600, 85.4500, "CDA Outer Channel tidal exchange zone"),
        ]

        # Extract 33 monitoring stations and 6 Nalabana sanctuary stations from water dataset
        water_file = DATA_BASE_DIR / "processed" / "chilika_water_processed.xlsx"
        if water_file.exists():
            wb_w = openpyxl.load_workbook(water_file, data_only=True)
            if "WATER_PROCESSED" in wb_w.sheetnames:
                ws_w = wb_w["WATER_PROCESSED"]
                for r in list(ws_w.iter_rows(values_only=True))[4:]:
                    if r and r[6]:
                        geo_str = str(r[6]).strip()
                        m = re.search(r"(.*?)\s*\(([0-9.]+),\s*([0-9.]+)\)", geo_str)
                        if m:
                            st_label = m.group(1).strip()
                            st_lat = float(m.group(2))
                            st_lon = float(m.group(3))
                            if not any(sl[0] == st_label for sl in spatial_locations):
                                spatial_locations.append((st_label, st_lat, st_lon, f"Water quality monitoring station ({st_label})"))

        existing_locs = {loc.label: loc for loc in dest.locations if loc.label}
        for label, lat, lon, desc in spatial_locations:
            if label not in existing_locs:
                loc = Location(
                    destination_id=dest.id,
                    label=label,
                    latitude=lat,
                    longitude=lon,
                    geojson=None,
                )
                self.db.add(loc)
                existing_locs[label] = loc
                self.report.locations_loaded += 1

        self.db.commit()
        self.db.refresh(dest)
        self._locations = {loc.label: loc for loc in dest.locations if loc.label}
        return dest

    # ── 2. Source Register & Datasets Ingestion ──────────────────────────────
    def load_sources_and_datasets(self) -> None:
        """
        Loads the 23 matched sources from SOURCE_ID_MAPPING and the 5 processed datasets.
        """
        sources_meta = [
            ("BIO-CDA-001", "Chilika Development Authority", "CDA Health Report Card 2021-22", "https://www.chilika.com/"),
            ("BIO-CDA-002", "Chilika Development Authority", "CDA Health Report Card 2023-24", "https://www.chilika.com/"),
            ("BIO-CDA-003", "ICAR-CIFRI & CDA", "Fish and Shellfish Diversity and Its Sustainable Management in Chilika Lake", "https://www.chilika.com/"),
            ("BIO-CDA-004", "Chilika Development Authority", "CDA Annual Report 2011-13", "https://www.chilika.com/"),
            ("BIO-CDA-006", "CDA & Wetland International", "CDA Seagrass / Biodiversity Report", "https://www.chilika.com/"),
            ("CH-COM-001", "FARD / Directorate of Fisheries", "Hand Book on Fisheries Statistics 2021-22", "https://fard.odisha.gov.in/"),
            ("CH-COM-002", "Directorate of Fisheries, Odisha", "Fisheries Administrative Report 2023-24", "https://fisheries.odisha.gov.in/"),
            ("CH-COM-003", "FISHFED Odisha", "FISHFED Annual Report 2021-22", "https://fishfed.odisha.gov.in/"),
            ("CH-COM-004", "Chilika Development Authority", "CDA Livelihood Support Guidelines", "https://www.chilika.com/"),
            ("CH-COM-011", "Directorate of Fisheries, Odisha", "FARD Training Report 2023-24", "https://fisheries.odisha.gov.in/"),
            ("CH-COM-012", "FARD / Directorate of Fisheries", "Hand Book on Fisheries Statistics 2012-13", "https://fard.odisha.gov.in/"),
            ("CH-FISH-002", "Directorate of Fisheries, Odisha", "Fisheries Statistics Handbook 2023-24", "https://fisheries.odisha.gov.in/"),
            ("CH-FISH-003", "ICAR-CIFRI", "CIFRI Bulletin on Chilika MSY & Sustainability", "https://cifri.icar.gov.in/"),
            ("CH-FISH-004", "Chilika Development Authority", "CDA Fisheries Monitoring Report 2021-22", "https://www.chilika.com/"),
            ("CH-FISH-005", "Directorate of Fisheries, Odisha", "District Fisheries Production Report 2023-24", "https://fisheries.odisha.gov.in/"),
            ("CH-TOUR-001", "Department of Tourism, Odisha", "Odisha Tourism Statistics 2023", "https://dot.odishatourism.gov.in/"),
            ("CH-TOUR-003", "Chilika Development Authority", "CDA Tourism & Boating Report 2023-24", "https://www.chilika.com/"),
            ("CH-CDA-TOUR-001", "Chilika Development Authority", "CDA Infrastructure Register 2023", "https://www.chilika.com/"),
            ("CH-CDA-TOUR-002", "Department of Tourism, Odisha", "OTDC Accommodation Directory 2023", "https://dot.odishatourism.gov.in/"),
            ("CDA-W-002", "Chilika Development Authority", "CDA Monthly Water Quality Data 2024", "https://www.chilika.com/"),
            ("CDA-W-003", "Chilika Development Authority", "CDA Nalabana Water Quality Data 2024", "https://www.chilika.com/"),
            ("OSPCB-ANNUAL-LAKES-2023", "Odisha State Pollution Control Board", "OSPCB Annual Lake Water Quality Report 2023", "https://ospcboard.org/"),
            ("OSPCB-WQI-MAY2025", "Odisha State Pollution Control Board", "OSPCB Water Quality Index May 2025", "https://ospcboard.org/"),
        ]

        for s_name, org, desc, url in sources_meta:
            source = self.source_repo.get_by_name(s_name)
            if not source:
                source = self.source_repo.create(
                    name=s_name,
                    organisation=org,
                    description=desc,
                    url=url,
                )
                self.report.sources_loaded += 1
            self._sources[s_name] = source

        datasets_meta = [
            ("CHL-TOUR-001", "chilika_tourism_processed.xlsx", "CH-TOUR-003", "2024-08-20", "Chilika Tourism Processed Dataset"),
            ("CHL-FIS-001", "chilika_fisheries_processed.xlsx", "CH-FISH-002", "2024-08-20", "Chilika Fisheries Processed Dataset"),
            ("CHL-WAT-001", "chilika_water_processed.xlsx", "CDA-W-002", "2024-08-20", "Chilika Water Processed Dataset"),
            ("CHL-BIO-002", "chilika_biodiversity_processed.xlsx", "BIO-CDA-001", "2024-08-20", "Chilika Biodiversity Processed Dataset"),
            ("CHL-COM-001", "chilika_community_processed.xlsx", "CH-COM-001", "2024-08-20", "Chilika Community Processed Dataset"),
        ]

        for ds_code, ds_name, primary_src, pub_date_str, desc in datasets_meta:
            src = self._sources.get(primary_src)
            if not src:
                continue

            stmt = select(Dataset).where(Dataset.name == ds_name, Dataset.source_id == src.id)
            ds = self.db.scalars(stmt).first()
            if not ds:
                p_date = date.fromisoformat(pub_date_str) if pub_date_str else None
                ds = self.dataset_repo.create(
                    name=ds_name,
                    source_id=src.id,
                    version="1.0",
                    publication_date=p_date,
                    description=f"{ds_code}: {desc}",
                )
                self.report.datasets_loaded += 1
            self._datasets[ds_code] = ds
            self._datasets[ds_name] = ds

    # ── 3. Granular Metric Registry Synchronization ──────────────────────────
    def load_safe_metric_definitions(self) -> None:
        """
        Creates granular MetricDefinition records required to distinguish water parameters,
        fisheries breakdowns, biodiversity indicators, community units, and tourism facilities.
        """
        granular_metrics = [
            # Water Quality
            ("water_fecal_coliform", "Fecal Coliform Concentration", "Water Quality", "MPN/100 mL", MetricDirection.LOWER_IS_BETTER, "Fecal Coliform concentration in water."),
            ("water_total_coliform", "Total Coliform Concentration", "Water Quality", "MPN/100 mL", MetricDirection.LOWER_IS_BETTER, "Total Coliform concentration in water."),
            ("water_ph", "Water pH Level", "Water Quality", "unitless", MetricDirection.NEUTRAL, "Water pH measurement."),
            ("water_dissolved_oxygen", "Dissolved Oxygen (DO)", "Water Quality", "mg/L", MetricDirection.HIGHER_IS_BETTER, "Dissolved oxygen concentration in water."),
            ("water_bod", "Biochemical Oxygen Demand (BOD)", "Water Quality", "mg/L", MetricDirection.LOWER_IS_BETTER, "Biochemical oxygen demand in water."),
            ("water_temperature", "Water Temperature", "Water Quality", "°C", MetricDirection.NEUTRAL, "Surface water temperature."),
            ("water_quality_parameter", "Water quality parameter", "Water Quality", "mg/L", MetricDirection.NEUTRAL, "CDA water quality parameters."),
            ("nalabana_water_quality", "Nalabana water quality", "Water Quality", "mg/L", MetricDirection.NEUTRAL, "Nalabana sanctuary water quality parameters."),
            ("lake_water_quality", "Lake water quality", "Water Quality", "mg/L", MetricDirection.NEUTRAL, "OSPCB annual lake water quality parameters."),
            ("water_quality_index", "Water quality index", "Water Quality", "index", MetricDirection.NEUTRAL, "OSPCB single-day water quality index."),
            # Fisheries
            ("fish_landings_total", "Total Fish Catch & Landings", "Fisheries", "metric ton (MT)", MetricDirection.NEUTRAL, "Total fish, prawn, and crab catch."),
            ("fish_landings_value", "Total Fish Landings Value", "Fisheries", "Million INR", MetricDirection.NEUTRAL, "Total economic value of fisheries landings."),
            ("fish_composition_prawn", "Fish Catch Composition - Prawn", "Fisheries", "%", MetricDirection.NEUTRAL, "Percentage composition of prawn in total catch."),
            ("fish_composition_crab", "Fish Catch Composition - Crab", "Fisheries", "%", MetricDirection.NEUTRAL, "Percentage composition of crab in total catch."),
            ("fish_composition_finfish", "Fish Catch Composition - Finfish", "Fisheries", "%", MetricDirection.NEUTRAL, "Percentage composition of finfish in total catch."),
            ("fish_production_annual", "Annual Fish Production Series", "Fisheries", "metric ton (MT)", MetricDirection.NEUTRAL, "District and lagoon annual fish production series."),
            ("fisheries_health_grade", "Fisheries Ecosystem Health Grade", "Fisheries", "ordinal scale", MetricDirection.NEUTRAL, "Report card health grade for fisheries."),
            ("commercial_species_diversity", "Desired Commercial Species Diversity", "Fisheries", "species", MetricDirection.HIGHER_IS_BETTER, "Target commercial finfish/shellfish species count."),
            ("sustainable_size_mullet", "Sustainable Size - Mullet (Mugil cephalus)", "Fisheries", "mm TL", MetricDirection.NEUTRAL, "Sustainable length threshold for Mullet."),
            ("sustainable_size_prawn", "Sustainable Size - Tiger Prawn (Penaeus monodon)", "Fisheries", "mm", MetricDirection.NEUTRAL, "Sustainable length threshold for Tiger Prawn."),
            ("sustainable_size_crab", "Sustainable Size - Crab (Scylla serrata)", "Fisheries", "mm", MetricDirection.NEUTRAL, "Sustainable carapace width threshold for Crab."),
            ("species_richness_finfish", "Finfish Species Richness", "Fisheries", "species", MetricDirection.HIGHER_IS_BETTER, "Total finfish species checklist count."),
            ("species_richness_shellfish", "Shellfish Species Richness", "Fisheries", "species", MetricDirection.HIGHER_IS_BETTER, "Total shellfish species checklist count."),
            ("near_threatened_finfish_species", "Near Threatened Finfish Species", "Fisheries", "species", MetricDirection.NEUTRAL, "Near Threatened finfish species count."),
            ("threatened_finfish_species", "Threatened Finfish Species Count", "Fisheries", "species", MetricDirection.LOWER_IS_BETTER, "Threatened (CR+EN+VU) finfish species count."),
            ("species_richness_not_evaluated", "Not Evaluated Finfish Species Count", "Fisheries", "species", MetricDirection.NEUTRAL, "Not evaluated species count."),
            ("maximum_sustainable_yield", "Maximum sustainable yield", "Fisheries", "tonnes/yr", MetricDirection.NEUTRAL, "CIFRI Maximum Sustainable Yield reference threshold."),
            ("fish_landings", "Fish landings", "Fisheries", "metric ton (MT)", MetricDirection.NEUTRAL, "Fish landings summary."),
            ("fish_production", "Fish production", "Fisheries", "metric ton (MT)", MetricDirection.NEUTRAL, "Fish production series."),
            ("district_fish_production", "District fish production", "Fisheries", "Million INR", MetricDirection.NEUTRAL, "District fish production and value."),
            ("fisheries_sustainability", "Fisheries sustainability", "Fisheries", "species", MetricDirection.NEUTRAL, "Fisheries sustainability targets."),
            ("fisheries_health_indicator", "Fisheries health indicator", "Fisheries", "ordinal scale", MetricDirection.NEUTRAL, "Fisheries health grades."),
            # Biodiversity
            ("biodiversity_msy_threshold", "Maximum Sustainable Yield Threshold", "Biodiversity", "tonnes/year", MetricDirection.NEUTRAL, "CDA biodiversity MSY threshold target."),
            ("ecosystem_health_grade", "Overall Ecosystem Health Grade", "Biodiversity", "ordinal scale (A, B+)", MetricDirection.NEUTRAL, "Overall ecosystem health letter grade."),
            ("bird_species_richness_study", "Avifauna Species Richness (Study Area)", "Biodiversity", "species", MetricDirection.HIGHER_IS_BETTER, "Total bird species count in/around study area."),
            ("birds_least_concern", "Birds - Least Concern (IUCN)", "Biodiversity", "species", MetricDirection.HIGHER_IS_BETTER, "Least Concern bird species count."),
            ("birds_near_threatened", "Birds - Near Threatened (IUCN)", "Biodiversity", "species", MetricDirection.NEUTRAL, "Near Threatened bird species count."),
            ("birds_vulnerable", "Birds - Vulnerable (IUCN)", "Biodiversity", "species", MetricDirection.LOWER_IS_BETTER, "Vulnerable bird species count."),
            ("birds_endangered", "Birds - Endangered (IUCN)", "Biodiversity", "species", MetricDirection.LOWER_IS_BETTER, "Endangered bird species count."),
            ("fishing_cat_population", "Fishing Cat Population Estimate", "Biodiversity", "individuals", MetricDirection.HIGHER_IS_BETTER, "Fishing cat population census count."),
            ("total_bird_census_count", "Total Bird Census Population Count", "Biodiversity", "individuals", MetricDirection.HIGHER_IS_BETTER, "Total annual bird census count."),
            ("floral_species_richness_angiosperm", "Flowering Plant Species Richness", "Biodiversity", "species", MetricDirection.HIGHER_IS_BETTER, "Total angiosperm plant species richness."),
            ("floral_species_richness_angiosperm_rprc", "Angiosperm Species (RPRC Record)", "Biodiversity", "species", MetricDirection.HIGHER_IS_BETTER, "Angiosperm species count (RPRC)."),
            ("phytoplankton_species_richness", "Phytoplankton Species Richness", "Biodiversity", "species", MetricDirection.HIGHER_IS_BETTER, "Phytoplankton species richness inventory."),
            ("nalabana_sanctuary_area", "Nalabana Protected Sanctuary Area", "Biodiversity", "square kilometer", MetricDirection.NEUTRAL, "Nalabana Bird Sanctuary core area extent."),
            ("endangered_species_indicator", "Endangered species indicator", "Biodiversity", "species", MetricDirection.LOWER_IS_BETTER, "Vulnerable and Endangered bird species counts."),
            ("species_richness", "Species richness", "Biodiversity", "species", MetricDirection.NEUTRAL, "Checklist monograph species richness accounts."),
            ("biodiversity_threshold", "Biodiversity threshold", "Biodiversity", "tonnes/year", MetricDirection.NEUTRAL, "Official CDA conservation threshold targets."),
            ("ecosystem_health_index", "Ecosystem health index", "Biodiversity", "ordinal scale (A, B+)", MetricDirection.NEUTRAL, "Overall ecosystem health letter grades from CDA index."),
            ("fish_species_richness", "Fish species richness", "Biodiversity", "species", MetricDirection.NEUTRAL, "Finfish species richness from scientific monograph inventories."),
            ("shellfish_species_richness", "Shellfish species richness", "Biodiversity", "species", MetricDirection.NEUTRAL, "Shellfish species richness from scientific monograph inventories."),
            ("biodiversity_population", "Biodiversity population", "Biodiversity", "individuals", MetricDirection.NEUTRAL, "Population count estimates."),
            ("bird_species_richness", "Bird species richness", "Biodiversity", "species", MetricDirection.NEUTRAL, "Avifauna species richness and IUCN conservation status counts."),
            ("floral_species_richness", "Floral species richness", "Biodiversity", "species", MetricDirection.NEUTRAL, "Flowering plant species richness records."),
            ("phytoplankton_richness", "Phytoplankton richness", "Biodiversity", "species", MetricDirection.NEUTRAL, "Phytoplankton species richness inventory."),
            ("protected_habitat_area", "Protected habitat area", "Biodiversity", "square kilometer", MetricDirection.NEUTRAL, "Protected habitat area extent."),
            # Community
            ("community_fisher_villages", "Fisher Villages Count", "Community", "villages", MetricDirection.NEUTRAL, "Number of fishermen villages in Chilika."),
            ("community_fisher_households", "Fishing Households Count", "Community", "households", MetricDirection.NEUTRAL, "Number of fishing households in Chilika."),
            ("community_fisher_population", "Total Fisher Population Count", "Community", "persons", MetricDirection.NEUTRAL, "Total fishermen population count."),
            ("community_fishers_solely_dependent", "Fishers Solely Dependent on Fishing", "Community", "persons", MetricDirection.NEUTRAL, "Number of fishers solely dependent on fishing."),
            ("community_fisher_income_per_capita", "Average Fisher Annual Per-Capita Income", "Community", "INR", MetricDirection.HIGHER_IS_BETTER, "Average annual per-capita income of active fishers."),
            ("community_pfcs_soft_loans", "PFCSs Provided Soft Loans", "Community", "societies", MetricDirection.NEUTRAL, "Number of Primary Fishermen Cooperative Societies provided soft loans."),
            ("community_loan_disbursement", "Total Soft Loan Amount Disbursed", "Community", "lakh INR", MetricDirection.NEUTRAL, "Total soft loan amount disbursed to PFCSs."),
            ("community_fishfed_pfcs_count", "FISHFED Primary Cooperative Societies Count", "Community", "societies", MetricDirection.NEUTRAL, "Number of PFCSs affiliated with FISHFED."),
            ("community_ifb_boxes_supplied", "Insulated Fish Boxes (IFB) Supplied", "Community", "boxes", MetricDirection.NEUTRAL, "Annual Insulated Fish Boxes supplied."),
            ("community_ifb_total_boxes_supplied", "Insulated Fish Boxes (IFB) Total Supplied", "Community", "boxes", MetricDirection.NEUTRAL, "Total Insulated Fish Boxes supplied under programme."),
            ("community_ifb_fishers_benefited", "Fishers Benefited by IFB", "Community", "persons", MetricDirection.NEUTRAL, "Number of traditional fishers benefited by IFB."),
            ("community_ifb_price_realization", "Higher Price Realization from IFB", "Community", "%", MetricDirection.HIGHER_IS_BETTER, "Percentage higher price realization due to IFB."),
            ("community_ifb_mpeda_subsidy", "MPEDA Subsidy for IFB Programme", "Community", "INR", MetricDirection.NEUTRAL, "MPEDA financial subsidy amount."),
            ("community_ifb_cda_subsidy", "CDA Subsidy for IFB Programme", "Community", "INR", MetricDirection.NEUTRAL, "CDA financial subsidy amount."),
            ("community_fishers_trained", "Fishers Trained (NETFISH-CDA)", "Community", "persons", MetricDirection.NEUTRAL, "Number of traditional fishers trained."),
            ("community_training_camps", "Training Camps Conducted", "Community", "camps", MetricDirection.NEUTRAL, "Number of training camps conducted."),
            ("community_training_villages", "Fishing Villages Covered by Training", "Community", "villages", MetricDirection.NEUTRAL, "Number of fishing villages covered by training."),
            ("community_training_pfcs_covered", "PFCSs Covered by Training", "Community", "societies", MetricDirection.NEUTRAL, "Number of PFCSs covered by training."),
            ("community_fisher_dependence", "Community fisher dependence", "Community", "villages", MetricDirection.NEUTRAL, "Fisher dependence indicators."),
            ("community_cooperative_coverage", "Community cooperative coverage", "Community", "societies", MetricDirection.NEUTRAL, "Cooperative coverage indicators."),
            ("community_livelihood_support", "Community livelihood support", "Community", "boxes", MetricDirection.NEUTRAL, "Livelihood support indicators."),
            ("community_training", "Community training", "Community", "persons", MetricDirection.NEUTRAL, "Community training participant counts."),
            # Tourism
            ("trained_boatmen_count", "Boatmen Trained in Safety & Skills", "Tourism", "persons", MetricDirection.NEUTRAL, "Boatmen who completed tourism and safety training."),
            ("cruise_vessels_count", "Day Cruise Vessels Count", "Tourism", "vessels", MetricDirection.NEUTRAL, "Number of operational day cruise vessels."),
            ("cruise_seating_capacity", "Seating Capacity per Day Cruise", "Tourism", "seats", MetricDirection.NEUTRAL, "Passenger seating capacity per cruise vessel."),
            ("houseboat_vessels_count", "Houseboat Vessels Count", "Tourism", "vessels", MetricDirection.NEUTRAL, "Number of operational luxury houseboats."),
            ("ferry_vessels_count", "Tourism & Access Ferry Vessels Count", "Tourism", "vessels", MetricDirection.NEUTRAL, "Number of passenger transport ferry vessels."),
            ("identified_tourist_centres_count", "Identified Tourist Centres Count", "Tourism", "centres", MetricDirection.NEUTRAL, "Number of designated tourism centres."),
            ("hotel_capacity_rooms", "Accommodation Room Capacity", "Tourism", "rooms", MetricDirection.NEUTRAL, "Total available accommodation rooms."),
            ("hotel_capacity_beds", "Accommodation Bed Capacity", "Tourism", "beds", MetricDirection.NEUTRAL, "Total available accommodation beds."),
            ("hotel_facilities_count", "Hotels & Facilities Count", "Tourism", "establishments", MetricDirection.NEUTRAL, "Total accommodation establishments."),
            ("hotel_occupancy_rate", "Annual Hotel Occupancy Rate", "Tourism", "%", MetricDirection.NEUTRAL, "Annual hotel occupancy percentage."),
            ("tourist_footfall_domestic", "Domestic Tourist Footfall", "Tourism", "visits", MetricDirection.NEUTRAL, "Annual domestic tourist visits."),
            ("tourist_footfall_foreign", "Foreign Tourist Footfall", "Tourism", "visits", MetricDirection.NEUTRAL, "Annual foreign tourist visits."),
            ("tourist_footfall_total", "Total Tourist Footfall", "Tourism", "visits", MetricDirection.NEUTRAL, "Total annual tourist footfall."),
            ("boating_activity", "Boating activity", "Tourism", "persons", MetricDirection.NEUTRAL, "Boating and skills participant counts."),
            ("tourism_facilities", "Tourism facilities", "Tourism", "cruises", MetricDirection.NEUTRAL, "Tourism physical facilities and capacity."),
            ("hotel_capacity", "Hotel capacity", "Tourism", "establishments", MetricDirection.NEUTRAL, "Accommodation capacity in Chilika area."),
            ("hotel_occupancy", "Hotel occupancy", "Tourism", "%", MetricDirection.NEUTRAL, "Hotel occupancy rate."),
            ("tourism_infrastructure", "Tourism infrastructure", "Tourism", "facilities", MetricDirection.NEUTRAL, "Tourism infrastructure and centres."),
            ("tourist_footfall", "Tourist footfall", "Tourism", "visits", MetricDirection.NEUTRAL, "Annual tourist footfall."),
        ]

        for code, name, cat, unit, direction, desc in granular_metrics:
            m = self.metric_repo.get_by_code_version(code, "1.0")
            if not m:
                m = self.metric_repo.create(
                    code=code,
                    version="1.0",
                    name=name,
                    category=cat,
                    unit=unit,
                    direction=direction,
                    description=desc,
                )
                self.report.metrics_loaded += 1
            self._metrics[code] = m

        for blocked_m in sorted(list(self._blocked_metrics)):
            self.report.unmapped_metrics.append(f"{blocked_m} (BLOCKED: excluded from MetricDefinition intake)")

    # ── 1. Destination & Locations Setup ─────────────────────────────────────
    def setup_destination_and_locations(self) -> Destination:
        """
        Creates or fetches the Chilika Destination and its spatial child Locations (52 total).
        """
        dest = self.dest_repo.get_by_name("Chilika")
        if not dest:
            dest = self.dest_repo.create(
                name="Chilika",
                description="Chilika Lagoon - Ramsar wetland of international importance (Site No. 229) and EcoTrace pilot destination.",
                country_code="IND",
                region="Odisha",
            )
            logger.info(f"Created Destination: {dest.name} (id={dest.id})")
        self._destination = dest

        # Core spatial locations from DESTINATION_GEOGRAPHY_MAPPING
        spatial_locations = [
            ("Satapada", 19.6756, 85.4412, "Outer channel tourism hub and dolphin watching base"),
            ("Barkul", 19.6897, 85.1952, "Central lagoon tourism hub and water sports base"),
            ("Rambha", 19.5167, 85.1000, "Southern lagoon tourism hub and island excursion base"),
            ("Nalabana", 19.6917, 85.2917, "Nalabana Bird Sanctuary island (protected area)"),
            ("Panthanivas Barkul", 19.6897, 85.1952, "OTDC accommodation facility at Barkul"),
            ("Panthanivas Rambha", 19.5167, 85.1000, "OTDC accommodation facility at Rambha"),
            ("Chilika - Khurda", 19.8000, 85.3000, "Khurda district administrative sector of Chilika"),
            ("Chilika - Puri", 19.7500, 85.5000, "Puri district administrative sector of Chilika"),
            ("Chilika - Ganjam", 19.5000, 85.1000, "Ganjam district administrative sector of Chilika"),
            ("Southern Zone", 19.5200, 85.1500, "CDA Southern ecological zone"),
            ("Central Zone", 19.6800, 85.2800, "CDA Central ecological zone"),
            ("Northern Zone", 19.8000, 85.3500, "CDA Northern ecological zone"),
            ("Outer Channel Zone", 19.6600, 85.4500, "CDA Outer Channel tidal exchange zone"),
        ]

        # Extract 33 monitoring stations and 6 Nalabana sanctuary stations from water dataset
        water_file = DATA_BASE_DIR / "processed" / "chilika_water_processed.xlsx"
        if water_file.exists():
            wb_w = openpyxl.load_workbook(water_file, data_only=True)
            if "WATER_PROCESSED" in wb_w.sheetnames:
                ws_w = wb_w["WATER_PROCESSED"]
                for r in list(ws_w.iter_rows(values_only=True))[4:]:
                    if r and r[6]:
                        geo_str = str(r[6]).strip()
                        m = re.search(r"(.*?)\s*\(([0-9.]+),\s*([0-9.]+)\)", geo_str)
                        if m:
                            st_label = m.group(1).strip()
                            st_lat = float(m.group(2))
                            st_lon = float(m.group(3))
                            if not any(sl[0] == st_label for sl in spatial_locations):
                                spatial_locations.append((st_label, st_lat, st_lon, f"Water quality monitoring station ({st_label})"))

        existing_locs = {loc.label: loc for loc in dest.locations if loc.label}
        for label, lat, lon, desc in spatial_locations:
            if label not in existing_locs:
                loc = Location(
                    destination_id=dest.id,
                    label=label,
                    latitude=lat,
                    longitude=lon,
                    geojson=None,
                )
                self.db.add(loc)
                existing_locs[label] = loc
                self.report.locations_loaded += 1

        self.db.commit()
        self.db.refresh(dest)
        self._locations = {loc.label: loc for loc in dest.locations if loc.label}
        return dest

    # ── 2. Source Register & Datasets Ingestion ──────────────────────────────
    def load_sources_and_datasets(self) -> None:
        """
        Loads the 23 matched sources from SOURCE_ID_MAPPING and the 5 processed datasets.
        """
        sources_meta = [
            ("BIO-CDA-001", "Chilika Development Authority", "CDA Health Report Card 2021-22", "https://www.chilika.com/"),
            ("BIO-CDA-002", "Chilika Development Authority", "CDA Health Report Card 2023-24", "https://www.chilika.com/"),
            ("BIO-CDA-003", "ICAR-CIFRI & CDA", "Fish and Shellfish Diversity and Its Sustainable Management in Chilika Lake", "https://www.chilika.com/"),
            ("BIO-CDA-004", "Chilika Development Authority", "CDA Annual Report 2011-13", "https://www.chilika.com/"),
            ("BIO-CDA-006", "CDA & Wetland International", "CDA Seagrass / Biodiversity Report", "https://www.chilika.com/"),
            ("CH-COM-001", "FARD / Directorate of Fisheries", "Hand Book on Fisheries Statistics 2021-22", "https://fard.odisha.gov.in/"),
            ("CH-COM-002", "Directorate of Fisheries, Odisha", "Fisheries Administrative Report 2023-24", "https://fisheries.odisha.gov.in/"),
            ("CH-COM-003", "FISHFED Odisha", "FISHFED Annual Report 2021-22", "https://fishfed.odisha.gov.in/"),
            ("CH-COM-004", "Chilika Development Authority", "CDA Livelihood Support Guidelines", "https://www.chilika.com/"),
            ("CH-COM-011", "Directorate of Fisheries, Odisha", "FARD Training Report 2023-24", "https://fisheries.odisha.gov.in/"),
            ("CH-COM-012", "FARD / Directorate of Fisheries", "Hand Book on Fisheries Statistics 2012-13", "https://fard.odisha.gov.in/"),
            ("CH-FISH-002", "Directorate of Fisheries, Odisha", "Fisheries Statistics Handbook 2023-24", "https://fisheries.odisha.gov.in/"),
            ("CH-FISH-003", "ICAR-CIFRI", "CIFRI Bulletin on Chilika MSY & Sustainability", "https://cifri.icar.gov.in/"),
            ("CH-FISH-004", "Chilika Development Authority", "CDA Fisheries Monitoring Report 2021-22", "https://www.chilika.com/"),
            ("CH-FISH-005", "Directorate of Fisheries, Odisha", "District Fisheries Production Report 2023-24", "https://fisheries.odisha.gov.in/"),
            ("CH-TOUR-001", "Department of Tourism, Odisha", "Odisha Tourism Statistics 2023", "https://dot.odishatourism.gov.in/"),
            ("CH-TOUR-003", "Chilika Development Authority", "CDA Tourism & Boating Report 2023-24", "https://www.chilika.com/"),
            ("CH-CDA-TOUR-001", "Chilika Development Authority", "CDA Infrastructure Register 2023", "https://www.chilika.com/"),
            ("CH-CDA-TOUR-002", "Department of Tourism, Odisha", "OTDC Accommodation Directory 2023", "https://dot.odishatourism.gov.in/"),
            ("CDA-W-002", "Chilika Development Authority", "CDA Monthly Water Quality Data 2024", "https://www.chilika.com/"),
            ("CDA-W-003", "Chilika Development Authority", "CDA Nalabana Water Quality Data 2024", "https://www.chilika.com/"),
            ("OSPCB-ANNUAL-LAKES-2023", "Odisha State Pollution Control Board", "OSPCB Annual Lake Water Quality Report 2023", "https://ospcboard.org/"),
            ("OSPCB-WQI-MAY2025", "Odisha State Pollution Control Board", "OSPCB Water Quality Index May 2025", "https://ospcboard.org/"),
        ]

        for s_name, org, desc, url in sources_meta:
            source = self.source_repo.get_by_name(s_name)
            if not source:
                source = self.source_repo.create(
                    name=s_name,
                    organisation=org,
                    description=desc,
                    url=url,
                )
                self.report.sources_loaded += 1
            self._sources[s_name] = source

        datasets_meta = [
            ("CHL-TOUR-001", "chilika_tourism_processed.xlsx", "CH-TOUR-003", "2024-08-20", "Chilika Tourism Processed Dataset"),
            ("CHL-FIS-001", "chilika_fisheries_processed.xlsx", "CH-FISH-002", "2024-08-20", "Chilika Fisheries Processed Dataset"),
            ("CHL-WAT-001", "chilika_water_processed.xlsx", "CDA-W-002", "2024-08-20", "Chilika Water Processed Dataset"),
            ("CHL-BIO-002", "chilika_biodiversity_processed.xlsx", "BIO-CDA-001", "2024-08-20", "Chilika Biodiversity Processed Dataset"),
            ("CHL-COM-001", "chilika_community_processed.xlsx", "CH-COM-001", "2024-08-20", "Chilika Community Processed Dataset"),
        ]

        for ds_code, ds_name, primary_src, pub_date_str, desc in datasets_meta:
            src = self._sources.get(primary_src)
            if not src:
                continue

            stmt = select(Dataset).where(Dataset.name == ds_name, Dataset.source_id == src.id)
            ds = self.db.scalars(stmt).first()
            if not ds:
                p_date = date.fromisoformat(pub_date_str) if pub_date_str else None
                ds = self.dataset_repo.create(
                    name=ds_name,
                    source_id=src.id,
                    version="1.0",
                    publication_date=p_date,
                    description=f"{ds_code}: {desc}",
                )
                self.report.datasets_loaded += 1
            self._datasets[ds_code] = ds
            self._datasets[ds_name] = ds

    # ── 3. Granular Metric Registry Synchronization ──────────────────────────
    def load_safe_metric_definitions(self) -> None:
        """
        Creates granular MetricDefinition records required to distinguish water parameters,
        fisheries breakdowns, biodiversity indicators, community units, and tourism facilities.
        """
        granular_metrics = [
            # Water Quality
            ("water_fecal_coliform", "Fecal Coliform Concentration", "Water Quality", "MPN/100 mL", MetricDirection.LOWER_IS_BETTER, "Fecal Coliform concentration in water."),
            ("water_total_coliform", "Total Coliform Concentration", "Water Quality", "MPN/100 mL", MetricDirection.LOWER_IS_BETTER, "Total Coliform concentration in water."),
            ("water_ph", "Water pH Level", "Water Quality", "unitless", MetricDirection.NEUTRAL, "Water pH measurement."),
            ("water_dissolved_oxygen", "Dissolved Oxygen (DO)", "Water Quality", "mg/L", MetricDirection.HIGHER_IS_BETTER, "Dissolved oxygen concentration in water."),
            ("water_bod", "Biochemical Oxygen Demand (BOD)", "Water Quality", "mg/L", MetricDirection.LOWER_IS_BETTER, "Biochemical oxygen demand in water."),
            ("water_temperature", "Water Temperature", "Water Quality", "°C", MetricDirection.NEUTRAL, "Surface water temperature."),
            ("water_quality_parameter", "Water quality parameter", "Water Quality", "mg/L", MetricDirection.NEUTRAL, "CDA water quality parameters."),
            ("nalabana_water_quality", "Nalabana water quality", "Water Quality", "mg/L", MetricDirection.NEUTRAL, "Nalabana sanctuary water quality parameters."),
            ("lake_water_quality", "Lake water quality", "Water Quality", "mg/L", MetricDirection.NEUTRAL, "OSPCB annual lake water quality parameters."),
            ("water_quality_index", "Water quality index", "Water Quality", "index", MetricDirection.NEUTRAL, "OSPCB single-day water quality index."),
            # Fisheries
            ("fish_landings_total", "Total Fish Catch & Landings", "Fisheries", "metric ton (MT)", MetricDirection.NEUTRAL, "Total fish, prawn, and crab catch."),
            ("fish_landings_value", "Total Fish Landings Value", "Fisheries", "Million INR", MetricDirection.NEUTRAL, "Total economic value of fisheries landings."),
            ("fish_composition_prawn", "Fish Catch Composition - Prawn", "Fisheries", "%", MetricDirection.NEUTRAL, "Percentage composition of prawn in total catch."),
            ("fish_composition_crab", "Fish Catch Composition - Crab", "Fisheries", "%", MetricDirection.NEUTRAL, "Percentage composition of crab in total catch."),
            ("fish_composition_finfish", "Fish Catch Composition - Finfish", "Fisheries", "%", MetricDirection.NEUTRAL, "Percentage composition of finfish in total catch."),
            ("fish_production_annual", "Annual Fish Production Series", "Fisheries", "metric ton (MT)", MetricDirection.NEUTRAL, "District and lagoon annual fish production series."),
            ("fisheries_health_grade", "Fisheries Ecosystem Health Grade", "Fisheries", "ordinal scale", MetricDirection.NEUTRAL, "Report card health grade for fisheries."),
            ("commercial_species_diversity", "Desired Commercial Species Diversity", "Fisheries", "species", MetricDirection.HIGHER_IS_BETTER, "Target commercial finfish/shellfish species count."),
            ("sustainable_size_mullet", "Sustainable Size - Mullet (Mugil cephalus)", "Fisheries", "mm TL", MetricDirection.NEUTRAL, "Sustainable length threshold for Mullet."),
            ("sustainable_size_prawn", "Sustainable Size - Tiger Prawn (Penaeus monodon)", "Fisheries", "mm", MetricDirection.NEUTRAL, "Sustainable length threshold for Tiger Prawn."),
            ("sustainable_size_crab", "Sustainable Size - Crab (Scylla serrata)", "Fisheries", "mm", MetricDirection.NEUTRAL, "Sustainable carapace width threshold for Crab."),
            ("species_richness_finfish", "Finfish Species Richness", "Fisheries", "species", MetricDirection.HIGHER_IS_BETTER, "Total finfish species checklist count."),
            ("species_richness_shellfish", "Shellfish Species Richness", "Fisheries", "species", MetricDirection.HIGHER_IS_BETTER, "Total shellfish species checklist count."),
            ("near_threatened_finfish_species", "Near Threatened Finfish Species", "Fisheries", "species", MetricDirection.NEUTRAL, "Near Threatened finfish species count."),
            ("threatened_finfish_species", "Threatened Finfish Species Count", "Fisheries", "species", MetricDirection.LOWER_IS_BETTER, "Threatened (CR+EN+VU) finfish species count."),
            ("species_richness_not_evaluated", "Not Evaluated Finfish Species Count", "Fisheries", "species", MetricDirection.NEUTRAL, "Not evaluated species count."),
            ("maximum_sustainable_yield", "Maximum sustainable yield", "Fisheries", "tonnes/yr", MetricDirection.NEUTRAL, "CIFRI Maximum Sustainable Yield reference threshold."),
            # Biodiversity
            ("biodiversity_msy_threshold", "Maximum Sustainable Yield Threshold", "Biodiversity", "tonnes/year", MetricDirection.NEUTRAL, "CDA biodiversity MSY threshold target."),
            ("ecosystem_health_grade", "Overall Ecosystem Health Grade", "Biodiversity", "ordinal scale (A, B+)", MetricDirection.NEUTRAL, "Overall ecosystem health letter grade."),
            ("bird_species_richness_study", "Avifauna Species Richness (Study Area)", "Biodiversity", "species", MetricDirection.HIGHER_IS_BETTER, "Total bird species count in/around study area."),
            ("birds_least_concern", "Birds - Least Concern (IUCN)", "Biodiversity", "species", MetricDirection.HIGHER_IS_BETTER, "Least Concern bird species count."),
            ("birds_near_threatened", "Birds - Near Threatened (IUCN)", "Biodiversity", "species", MetricDirection.NEUTRAL, "Near Threatened bird species count."),
            ("birds_vulnerable", "Birds - Vulnerable (IUCN)", "Biodiversity", "species", MetricDirection.LOWER_IS_BETTER, "Vulnerable bird species count."),
            ("birds_endangered", "Birds - Endangered (IUCN)", "Biodiversity", "species", MetricDirection.LOWER_IS_BETTER, "Endangered bird species count."),
            ("fishing_cat_population", "Fishing Cat Population Estimate", "Biodiversity", "individuals", MetricDirection.HIGHER_IS_BETTER, "Fishing cat population census count."),
            ("total_bird_census_count", "Total Bird Census Population Count", "Biodiversity", "individuals", MetricDirection.HIGHER_IS_BETTER, "Total annual bird census count."),
            ("floral_species_richness_angiosperm", "Flowering Plant Species Richness", "Biodiversity", "species", MetricDirection.HIGHER_IS_BETTER, "Total angiosperm plant species richness."),
            ("floral_species_richness_angiosperm_rprc", "Angiosperm Species (RPRC Record)", "Biodiversity", "species", MetricDirection.HIGHER_IS_BETTER, "Angiosperm species count (RPRC)."),
            ("phytoplankton_species_richness", "Phytoplankton Species Richness", "Biodiversity", "species", MetricDirection.HIGHER_IS_BETTER, "Phytoplankton species richness inventory."),
            ("nalabana_sanctuary_area", "Nalabana Protected Sanctuary Area", "Biodiversity", "square kilometer", MetricDirection.NEUTRAL, "Nalabana Bird Sanctuary core area extent."),
            ("endangered_species_indicator", "Endangered species indicator", "Biodiversity", "species", MetricDirection.LOWER_IS_BETTER, "Vulnerable and Endangered bird species counts."),
            # Community
            ("community_fisher_villages", "Fisher Villages Count", "Community", "villages", MetricDirection.NEUTRAL, "Number of fishermen villages in Chilika."),
            ("community_fisher_households", "Fishing Households Count", "Community", "households", MetricDirection.NEUTRAL, "Number of fishing households in Chilika."),
            ("community_fisher_population", "Total Fisher Population Count", "Community", "persons", MetricDirection.NEUTRAL, "Total fishermen population count."),
            ("community_fishers_solely_dependent", "Fishers Solely Dependent on Fishing", "Community", "persons", MetricDirection.NEUTRAL, "Number of fishers solely dependent on fishing."),
            ("community_fisher_income_per_capita", "Average Fisher Annual Per-Capita Income", "Community", "INR", MetricDirection.HIGHER_IS_BETTER, "Average annual per-capita income of active fishers."),
            ("community_pfcs_soft_loans", "PFCSs Provided Soft Loans", "Community", "societies", MetricDirection.NEUTRAL, "Number of Primary Fishermen Cooperative Societies provided soft loans."),
            ("community_loan_disbursement", "Total Soft Loan Amount Disbursed", "Community", "lakh INR", MetricDirection.NEUTRAL, "Total soft loan amount disbursed to PFCSs."),
            ("community_fishfed_pfcs_count", "FISHFED Primary Cooperative Societies Count", "Community", "societies", MetricDirection.NEUTRAL, "Number of PFCSs affiliated with FISHFED."),
            ("community_ifb_boxes_supplied", "Insulated Fish Boxes (IFB) Supplied", "Community", "boxes", MetricDirection.NEUTRAL, "Annual Insulated Fish Boxes supplied."),
            ("community_ifb_total_boxes_supplied", "Insulated Fish Boxes (IFB) Total Supplied", "Community", "boxes", MetricDirection.NEUTRAL, "Total Insulated Fish Boxes supplied under programme."),
            ("community_ifb_fishers_benefited", "Fishers Benefited by IFB", "Community", "persons", MetricDirection.NEUTRAL, "Number of traditional fishers benefited by IFB."),
            ("community_ifb_price_realization", "Higher Price Realization from IFB", "Community", "%", MetricDirection.HIGHER_IS_BETTER, "Percentage higher price realization due to IFB."),
            ("community_ifb_mpeda_subsidy", "MPEDA Subsidy for IFB Programme", "Community", "INR", MetricDirection.NEUTRAL, "MPEDA financial subsidy amount."),
            ("community_ifb_cda_subsidy", "CDA Subsidy for IFB Programme", "Community", "INR", MetricDirection.NEUTRAL, "CDA financial subsidy amount."),
            ("community_fishers_trained", "Fishers Trained (NETFISH-CDA)", "Community", "persons", MetricDirection.NEUTRAL, "Number of traditional fishers trained."),
            ("community_training_camps", "Training Camps Conducted", "Community", "camps", MetricDirection.NEUTRAL, "Number of training camps conducted."),
            ("community_training_villages", "Fishing Villages Covered by Training", "Community", "villages", MetricDirection.NEUTRAL, "Number of fishing villages covered by training."),
            ("community_training_pfcs_covered", "PFCSs Covered by Training", "Community", "societies", MetricDirection.NEUTRAL, "Number of PFCSs covered by training."),
            # Tourism
            ("trained_boatmen_count", "Boatmen Trained in Safety & Skills", "Tourism", "persons", MetricDirection.NEUTRAL, "Boatmen who completed tourism and safety training."),
            ("cruise_vessels_count", "Day Cruise Vessels Count", "Tourism", "vessels", MetricDirection.NEUTRAL, "Number of operational day cruise vessels."),
            ("cruise_seating_capacity", "Seating Capacity per Day Cruise", "Tourism", "seats", MetricDirection.NEUTRAL, "Passenger seating capacity per cruise vessel."),
            ("houseboat_vessels_count", "Houseboat Vessels Count", "Tourism", "vessels", MetricDirection.NEUTRAL, "Number of operational luxury houseboats."),
            ("ferry_vessels_count", "Tourism & Access Ferry Vessels Count", "Tourism", "vessels", MetricDirection.NEUTRAL, "Number of passenger transport ferry vessels."),
            ("identified_tourist_centres_count", "Identified Tourist Centres Count", "Tourism", "centres", MetricDirection.NEUTRAL, "Number of designated tourism centres."),
            ("hotel_capacity_rooms", "Accommodation Room Capacity", "Tourism", "rooms", MetricDirection.NEUTRAL, "Total available accommodation rooms."),
            ("hotel_capacity_beds", "Accommodation Bed Capacity", "Tourism", "beds", MetricDirection.NEUTRAL, "Total available accommodation beds."),
            ("hotel_facilities_count", "Hotels & Facilities Count", "Tourism", "establishments", MetricDirection.NEUTRAL, "Total accommodation establishments."),
            ("hotel_occupancy_rate", "Annual Hotel Occupancy Rate", "Tourism", "%", MetricDirection.NEUTRAL, "Annual hotel occupancy percentage."),
            ("tourist_footfall_domestic", "Domestic Tourist Footfall", "Tourism", "visits", MetricDirection.NEUTRAL, "Annual domestic tourist visits."),
            ("tourist_footfall_foreign", "Foreign Tourist Footfall", "Tourism", "visits", MetricDirection.NEUTRAL, "Annual foreign tourist visits."),
            ("tourist_footfall_total", "Total Tourist Footfall", "Tourism", "visits", MetricDirection.NEUTRAL, "Total annual tourist footfall."),
            ("boating_activity", "Boating activity", "Tourism", "persons", MetricDirection.NEUTRAL, "Boating and skills participant counts."),
            ("tourism_facilities", "Tourism facilities", "Tourism", "cruises", MetricDirection.NEUTRAL, "Tourism physical facilities and capacity."),
            ("hotel_capacity", "Hotel capacity", "Tourism", "establishments", MetricDirection.NEUTRAL, "Accommodation capacity in Chilika area."),
            ("tourist_footfall", "Tourist footfall", "Tourism", "visits", MetricDirection.NEUTRAL, "Annual tourist footfall."),
        ]

        for code, name, cat, unit, direction, desc in granular_metrics:
            m = self.metric_repo.get_by_code_version(code, "1.0")
            if not m:
                m = self.metric_repo.create(
                    code=code,
                    version="1.0",
                    name=name,
                    category=cat,
                    unit=unit,
                    direction=direction,
                    description=desc,
                )
                self.report.metrics_loaded += 1
            self._metrics[code] = m

        for blocked_m in sorted(list(self._blocked_metrics)):
            self.report.unmapped_metrics.append(f"{blocked_m} (BLOCKED: excluded from MetricDefinition intake)")

    # ── 4. Observation & Evidence Ingestion ──────────────────────────────────
    def ingest_processed_file(self, filename: str, ds_code: str, sheet_name: str) -> None:
        """
        Reads and ingests a single processed Excel file.
        Maps spatial locations and granular metrics cleanly without collapsing distinct records.
        """
        file_path = DATA_BASE_DIR / "processed" / filename
        if not file_path.exists():
            err_msg = f"Processed file not found: {file_path}"
            self.report.validation_errors.append(err_msg)
            logger.error(err_msg)
            return

        self.report.files_processed.append(filename)
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            err_msg = f"Sheet {sheet_name} not found in {filename}"
            self.report.validation_errors.append(err_msg)
            return

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        header_idx = -1
        for idx, r in enumerate(rows):
            if r and "dataset_id" in [str(c).strip() if c else "" for c in r]:
                header_idx = idx
                break

        if header_idx == -1:
            self.report.validation_errors.append(f"Header row not found in {filename}")
            return

        header = [str(c).strip() if c else "" for c in rows[header_idx]]
        data_rows = [dict(zip(header, r)) for r in rows[header_idx + 1:] if r and r[0] is not None]

        dataset = self._datasets.get(ds_code)
        if not dataset:
            self.report.validation_errors.append(f"Dataset {ds_code} not initialized in database")
            return

        for rec in data_rows:
            self.report.records_read += 1
            obs_id = str(rec.get("observation_id", "")).strip()
            metric_key = str(rec.get("regenledger_metric", "")).strip()
            rec_type = str(rec.get("record_type", "")).strip()
            raw_val = rec.get("value")
            raw_unit = str(rec.get("unit", "")).strip()
            raw_geo = str(rec.get("geography", "")).strip()
            raw_status = str(rec.get("data_status", "")).strip()
            raw_notes = str(rec.get("notes", "")).strip() if rec.get("notes") else ""
            source_id_str = str(rec.get("source_id", "")).strip()

            # ── Check for blocked items ──
            if obs_id in self._blocked_obs_ids or metric_key in self._blocked_metrics or rec_type == "needs_review":
                self.report.records_blocked += 1
                continue

            # ── Resolve Granular Metric & Location ──
            metric_code = self.resolve_metric_code(rec, filename)
            metric_def = self._metrics.get(metric_code) or self._metrics.get(metric_key)
            if not metric_def:
                self.report.records_skipped += 1
                continue

            loc_obj = self.resolve_location(raw_geo)
            loc_id = loc_obj.id if loc_obj else None

            # ── Parse Temporal Period ──
            period = self.parse_temporal_period(rec.get("year"))
            if not period:
                self.report.records_skipped += 1
                self.report.validation_errors.append(f"Unparseable date period '{rec.get('year')}' in {obs_id}")
                continue

            period_start, period_end = period

            # ── Parse Values ──
            orig_value: Optional[float] = None
            norm_value: Optional[float] = None
            if raw_val is not None:
                try:
                    if isinstance(raw_val, (int, float)):
                        orig_value = float(raw_val)
                        norm_value = float(raw_val)
                    else:
                        orig_value = float(str(raw_val).strip())
                        norm_value = orig_value
                except (ValueError, TypeError):
                    orig_value = None
                    norm_value = None

            # ── Status & Confidence Mapping ──
            obs_status = ObservationStatus.VERIFIED
            if "PROVISIONAL" in raw_status or "THRESHOLD" in raw_status:
                obs_status = ObservationStatus.RAW
            elif "REFERENCE" in raw_status:
                obs_status = ObservationStatus.VERIFIED

            confidence = ConfidenceLevel.HIGH
            if "HISTORICAL" in raw_status:
                confidence = ConfidenceLevel.MEDIUM
            elif "PROVISIONAL" in raw_status:
                confidence = ConfidenceLevel.LOW

            dest_spec = DestinationSpecificity.DIRECT
            if "statewide" in raw_geo.lower() or "STATEWIDE" in raw_status:
                dest_spec = DestinationSpecificity.REGIONAL

            # ── Structured Notes & Provenance Trace ──
            notes_payload = (
                f"observation_id: {obs_id} | "
                f"raw_indicator: {rec.get('indicator')} | "
                f"raw_value: {raw_val} | "
                f"raw_unit: {raw_unit} | "
                f"raw_geography: {raw_geo} | "
                f"source_id: {source_id_str} | "
                f"data_status: {raw_status} | "
                f"notes: {raw_notes}"
            )

            # ── Natural Key Collision & Idempotency Check ──
            # uq_observation_natural_key: (destination_id, location_id, metric_definition_id, dataset_id, period_start, period_end)
            stmt = select(Observation).where(
                Observation.destination_id == self._destination.id,
                Observation.location_id == loc_id,
                Observation.metric_definition_id == metric_def.id,
                Observation.dataset_id == dataset.id,
                Observation.period_start == period_start,
                Observation.period_end == period_end,
            )
            existing_obs = self.db.scalars(stmt).first()

            if existing_obs:
                self.report.duplicate_preventions += 1
                if existing_obs.original_value is None and orig_value is not None:
                    existing_obs.original_value = orig_value
                    existing_obs.normalized_value = norm_value
                    existing_obs.notes = notes_payload
                    self.db.add(existing_obs)
                    self.db.commit()
                    self.report.records_updated += 1
                continue

            # ── Create Observation ──
            obs = Observation(
                destination_id=self._destination.id,
                location_id=loc_id,
                metric_definition_id=metric_def.id,
                dataset_id=dataset.id,
                period_start=period_start,
                period_end=period_end,
                original_value=orig_value,
                normalized_value=norm_value,
                status=obs_status,
                confidence=confidence,
                destination_specificity=dest_spec,
                methodology=f"Indicator: {rec.get('indicator')} | Domain: {rec.get('taxonomic_group') or rec.get('community_domain') or rec.get('fisheries_domain') or rec.get('tourism_domain') or rec.get('water_domain') or ''} | Type: {rec_type}",
                assumptions=raw_status,
                notes=notes_payload,
            )
            self.db.add(obs)
            self.db.commit()
            self.db.refresh(obs)
            self.report.records_inserted += 1

            # ── Evidence Creation (Strict: Only when verifiable source metadata exists) ──
            source_obj = self._sources.get(source_id_str)
            if source_obj and raw_notes and ("source_page" in raw_notes or "source_record_id" in raw_notes or "Table" in raw_notes or "raw_file" in raw_notes):
                ev = Evidence(
                    observation_id=obs.id,
                    source_id=source_obj.id,
                    dataset_id=dataset.id,
                    evidence_type=EvidenceType.DOCUMENT,
                    reference_url=source_obj.url,
                    raw_excerpt=f"[{obs_id}] {rec.get('indicator')}: {raw_val} {raw_unit} ({rec.get('year')})",
                    notes=raw_notes,
                )
                self.db.add(ev)
                self.db.commit()
                self.report.evidence_created += 1
            else:
                self.report.provenance_gaps.append(
                    f"Observation {obs_id} (source={source_id_str}): Evidence ID unassigned (P2 Partial provenance preserved)"
                )

    # ── Targeted Cleanup Helper ──────────────────────────────────────────────
    def clear_chilika_observations(self) -> int:
        """
        Safely removes existing Chilika observations and evidence created by pilot runs.
        Leaves all sources, datasets, destinations, locations, and other destinations untouched.
        """
        dest = self.dest_repo.get_by_name("Chilika")
        if not dest:
            return 0

        # Delete observations (Evidence rows cascade-delete automatically)
        chilika_obs_ids = [
            o.id for o in self.db.scalars(
                select(Observation).where(Observation.destination_id == dest.id)
            ).all()
        ]
        if chilika_obs_ids:
            # Delete evidence explicitly just in case
            self.db.execute(delete(Evidence).where(Evidence.observation_id.in_(chilika_obs_ids)))
            self.db.execute(delete(Observation).where(Observation.id.in_(chilika_obs_ids)))
            self.db.commit()
            logger.info(f"Cleared {len(chilika_obs_ids)} existing Chilika observations and associated evidence.")
            return len(chilika_obs_ids)
        return 0

    # ── 5. Full Orchestration Runner ─────────────────────────────────────────
    def run(self, clean_existing_chilika: bool = False) -> IngestionReport:
        """
        Executes the full Chilika pilot ingestion pipeline in sequential, idempotent steps.
        """
        logger.info("Starting Chilika Pilot Ingestion Pipeline...")
        self.setup_destination_and_locations()
        self.load_sources_and_datasets()
        self.load_safe_metric_definitions()

        if clean_existing_chilika:
            self.clear_chilika_observations()

        files_to_ingest = [
            ("chilika_tourism_processed.xlsx", "CHL-TOUR-001", "TOURISM_PROCESSED"),
            ("chilika_fisheries_processed.xlsx", "CHL-FIS-001", "FISHERIES_PROCESSED"),
            ("chilika_water_processed.xlsx", "CHL-WAT-001", "WATER_PROCESSED"),
            ("chilika_biodiversity_processed.xlsx", "CHL-BIO-002", "BIODIVERSITY_PROCESSED"),
            ("chilika_community_processed.xlsx", "CHL-COM-001", "COMMUNITY_PROCESSED"),
        ]

        for fname, dscode, sname in files_to_ingest:
            self.ingest_processed_file(fname, dscode, sname)

        logger.info(f"Chilika Ingestion Completed. Inserted: {self.report.records_inserted}, Blocked: {self.report.records_blocked}, Deduped: {self.report.duplicate_preventions}")
        return self.report


def run_chilika_ingestion(clean_existing: bool = False) -> IngestionReport:
    """CLI/Helper entry point to execute the ingestion pipeline."""
    db = SessionLocal()
    try:
        service = ChilikaIngestionService(db)
        report = service.run(clean_existing_chilika=clean_existing)
        return report
    finally:
        db.close()


if __name__ == "__main__":
    rep = run_chilika_ingestion()
    print("\n" + "=" * 70)
    print("CHILIKA PILOT INGESTION EXECUTION REPORT")
    print("=" * 70)
    print(f"Files Processed:        {len(rep.files_processed)} files ({', '.join(rep.files_processed)})")
    print(f"Records Read:           {rep.records_read}")
    print(f"Records Inserted:       {rep.records_inserted}")
    print(f"Records Updated:        {rep.records_updated}")
    print(f"Records Skipped:        {rep.records_skipped}")
    print(f"Records Blocked:        {rep.records_blocked}")
    print(f"Duplicate Preventions:  {rep.duplicate_preventions}")
    print(f"Evidence Items Created: {rep.evidence_created}")
    print(f"Sources Loaded:         {rep.sources_loaded}")
    print(f"Datasets Loaded:        {rep.datasets_loaded}")
    print(f"Metrics Loaded:         {rep.metrics_loaded}")
    print(f"Locations Loaded:       {rep.locations_loaded}")
    print(f"Unmapped Metrics:       {len(rep.unmapped_metrics)}")
    print(f"Provenance Gaps (P2):   {len(rep.provenance_gaps)}")
    print(f"Validation Errors:      {len(rep.validation_errors)}")
    print("=" * 70 + "\n")
