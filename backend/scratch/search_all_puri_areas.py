import openpyxl
from pathlib import Path

p_dir = Path("c:/S21_new/backend/REGENLEDGER_DATA (1)/REGENLEDGER_DATA_PURI_UPDATED")
for fpath in p_dir.glob("**/*.xlsx"):
    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
        for sname in wb.sheetnames:
            ws = wb[sname]
            for r in ws.iter_rows(values_only=True):
                r_str = str(r).lower()
                if "shamuka" in r_str or "sq.km" in r_str or "sq km" in r_str or "sqkm" in r_str or "area" in r_str:
                    if any(w in r_str for w in ["municipal", "ulb", "km", "ha", "acre"]):
                        print(f"[{fpath.name} | {sname}] {r[:7]}")
    except Exception as e:
        pass
