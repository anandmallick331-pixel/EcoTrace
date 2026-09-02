import openpyxl
from pathlib import Path

print("=== 1. SEARCH AREA IN KONARK PACKAGE ===")
k_dir = Path("c:/S21_new/backend/REGENLEDGER_DATA (1)/REGENLEDGER_DATA_KONARK_UPDATED")
for fpath in [
    k_dir / "raw/01_RAW/KONARK_COMMUNITY_RAW.xlsx",
    k_dir / "raw/01_RAW/KONARK_GIS_RAW.xlsx",
    k_dir / "raw/01_RAW/KONARK_WASTE_RAW.xlsx",
    k_dir / "framework/KONARK_GEOGRAPHIC_ATTRIBUTION_RULES.xlsx",
    k_dir / "framework/KONARK_METRIC_STATUS_TABLE.xlsx",
    k_dir / "s21_ready/06_OBSERVATIONS.xlsx"
]:
    if fpath.exists():
        wb = openpyxl.load_workbook(fpath, data_only=True)
        print(f"\n--- {fpath.name} ---")
        ws = wb.active
        for r in ws.iter_rows(values_only=True):
            r_str = str(r).lower()
            if any(w in r_str for w in ["area", "sqkm", "sq km", "hectare", "boundary", "nac", "extent", "sq.km", "km2"]):
                print(r[:8])

print("\n=== 2. SEARCH AREA IN PURI PACKAGE ===")
p_dir = Path("c:/S21_new/backend/REGENLEDGER_DATA (1)/REGENLEDGER_DATA_PURI_UPDATED")
for fpath in [
    p_dir / "PROCESSED_DATA/PURI_COMMUNITY_PROCESSED.xlsx",
    p_dir / "PROCESSED_DATA/PURI_GIS_PROCESSED.xlsx",
    p_dir / "PROCESSED_DATA/PURI_WASTE_PROCESSED.xlsx",
    p_dir / "metadata/PURI_CANONICAL_OBSERVATIONS.xlsx",
    p_dir / "framework/PURI_METRIC_STATUS_TABLE.xlsx"
]:
    if fpath.exists():
        wb = openpyxl.load_workbook(fpath, data_only=True)
        print(f"\n--- {fpath.name} ---")
        ws = wb.active
        for r in ws.iter_rows(values_only=True):
            r_str = str(r).lower()
            if any(w in r_str for w in ["area", "sqkm", "sq km", "hectare", "municipality", "extent", "sq.km", "km2"]):
                print(r[:8])
