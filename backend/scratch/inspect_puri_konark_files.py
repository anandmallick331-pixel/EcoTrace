import openpyxl
from pathlib import Path

print("=== INSPECT PURI GIS / AREA / WASTE PROCESSED ===")
puri_dir = Path("c:/S21_new/backend/REGENLEDGER_DATA (1)/REGENLEDGER_DATA_PURI_UPDATED/PROCESSED_DATA")

for fname in ["PURI_GIS_PROCESSED.xlsx", "PURI_WASTE_PROCESSED.xlsx", "PURI_COMMUNITY_PROCESSED.xlsx", "PURI_ENVIRONMENT_PROCESSED.xlsx"]:
    fpath = puri_dir / fname
    if fpath.exists():
        wb = openpyxl.load_workbook(fpath)
        print(f"\n--- {fname} (Sheets: {wb.sheetnames}) ---")
        ws = wb.active
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            if idx < 10 and any(c is not None for c in row):
                print(row[:8])

print("\n=== INSPECT KONARK RAW DATA DIRECTORY ===")
konark_dir = Path("c:/S21_new/backend/REGENLEDGER_DATA (1)/REGENLEDGER_DATA_KONARK_UPDATED")
if not konark_dir.exists():
    konark_dir = Path("c:/S21_new/backend/REGENLEDGER_DATA (1)/Konark")
if not konark_dir.exists():
    konark_dir = Path("c:/S21_new/REGENLEDGER_DATA (1)/REGENLEDGER_DATA_KONARK_UPDATED")

print("Konark dir exists:", konark_dir.exists())
if konark_dir.exists():
    for f in konark_dir.glob("**/*.xlsx"):
        print(f)
