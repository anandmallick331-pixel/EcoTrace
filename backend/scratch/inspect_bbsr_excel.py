import openpyxl
from pathlib import Path

data_dir = Path("c:/S21_new/REGENLEDGER_DATA (1)/REGENLEDGER_DATA_BBSR_UPDATED/s21_ready")
if not data_dir.exists():
    data_dir = Path("c:/S21_new/REGENLEDGER_DATA (1)/Bhubaneswar/s21_ready")

print("Checking data_dir:", data_dir)
wb_met = openpyxl.load_workbook(data_dir / "03_METRIC_DEFINITIONS.xlsx")
ws_met = wb_met.active
print("--- ALL METRIC DEFINITIONS IN BBSR PACKAGE ---")
for r in ws_met.iter_rows(values_only=True):
    if r and any(c is not None for c in r):
        print(r[:5])

wb_obs = openpyxl.load_workbook(data_dir / "06_OBSERVATIONS.xlsx")
ws_obs = wb_obs.active
print("\n--- ALL OBSERVATIONS IN BBSR PACKAGE ---")
for r in ws_obs.iter_rows(values_only=True):
    if r and any(c is not None for c in r):
        if "pop" in str(r).lower() or "census" in str(r).lower() or "comm" in str(r).lower():
            print(r[:8])
